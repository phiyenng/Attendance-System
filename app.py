from flask import Flask, render_template, request, jsonify, send_file
from markupsafe import Markup
import pandas as pd
import openpyxl
import calendar
from datetime import datetime, timedelta, date
import os
import json
from werkzeug.utils import secure_filename
import numpy as np
from io import BytesIO
import re
import os

app = Flask(__name__)
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'your-secret-key-here-change-in-production')
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024

# Files path
TEMP_SIGNINOUT_PATH = os.path.join(app.config['UPLOAD_FOLDER'], 'temp_signinout.xlsx')
TEMP_APPLY_PATH = os.path.join(app.config['UPLOAD_FOLDER'], 'temp_apply.xlsx')
TEMP_OTLIEU_PATH = os.path.join(app.config['UPLOAD_FOLDER'], 'temp_otlieu.xlsx')
EMPLOYEE_LIST_PATH = os.path.join(app.config['UPLOAD_FOLDER'], 'employee_list.csv')
RULES_XLSX_PATH = os.path.join(app.config['UPLOAD_FOLDER'], 'rules.xlsx')

# Ensure upload folder exists
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# Global variables to store data with caching
sign_in_out_data = None
apply_data = None
ot_lieu_data = None
employee_list_df = None
rules = None
global lieu_followup_df
lieu_followup_df = None

# Cache for expensive calculations
_attendance_report_cache = {}
_total_attendance_cache = {}
_cache_timeout = 300  # 5 min

# Memory optimization settings
pd.set_option('mode.copy_on_write', True)  # Reduce memory usage in pandas operations

def _get_cache_key():
    """Generate cache key based on file modification times"""
    files = [TEMP_SIGNINOUT_PATH, TEMP_APPLY_PATH, TEMP_OTLIEU_PATH, EMPLOYEE_LIST_PATH, RULES_XLSX_PATH]
    mod_times = []
    for file_path in files:
        if os.path.exists(file_path):
            mod_times.append(str(int(os.path.getmtime(file_path))))
        else:
            mod_times.append("0")
    return "_".join(mod_times)

def _clear_cache():
    """Clear all cached data"""
    global _attendance_report_cache, _total_attendance_cache
    _attendance_report_cache.clear()
    _total_attendance_cache.clear()

# ==========================
# EMPLOYEE LIST
# ==========================
if os.path.exists(EMPLOYEE_LIST_PATH):
    try:
        employee_list_df = pd.read_csv(EMPLOYEE_LIST_PATH, dtype={'ID Number': str})
        if 'Dept' not in employee_list_df.columns:
            employee_list_df['Dept'] = ''
        if 'Internship' not in employee_list_df.columns:
            employee_list_df['Internship'] = ''
    except Exception:
        employee_list_df = pd.DataFrame(columns=["STT", "Name", "ID Number", "Dept", "Internship"])
else:
    employee_list_df = pd.DataFrame(columns=["STT", "Name", "ID Number", "Dept", "Internship"])

# ==========================
# RULES 
# ==========================
if os.path.exists(RULES_XLSX_PATH):
    try:
        rules = pd.read_excel(RULES_XLSX_PATH)
    except Exception as e:
        print(f"Error loading rules: {e}")
        rules = pd.DataFrame()
else:
    rules = pd.DataFrame(columns=['Holiday Date in This Year', 'Special Work Day'])
    try:
        rules.to_excel(RULES_XLSX_PATH, index=False)
        print(f"Created default rules file: {RULES_XLSX_PATH}")
    except Exception as e:
        print(f"Error creating rules file: {e}")
        rules = pd.DataFrame()

# Application Data - Translate Headers
APPLY_HEADER_MAP = {
    '申请人工号': 'Emp ID',
    '申请人': 'Emp Name',
    '提交人工号': 'Submit ID',
    '提交人': 'Submit Name',
    '申请时间': 'Apply Date',
    '起始时间': 'Start Date',
    '申请类型': 'Application Type',
    '终止时间': 'End Date',
    '申请说明': 'Note',
    '审批结果': 'Approve Result',
}

# =======================
# APPLY DATA
# =======================
def map_leave_type(note):
    if pd.isna(note):
        return ''
    note = str(note)
    if re.search(r'事假|Leave|unpaid', note, re.I):
        return 'Unpaid'
    if re.search(r'年休假|Annual', note, re.I):
        return 'Annual'
    if re.search(r'产|婚|育|丧|welfare', note, re.I):
        return 'Welfare'
    if re.search(r'病假|sick', note, re.I):
        return 'Sick'
    return note

def translate_apply_headers(df):
    def normalize(col):
        return re.sub(r"[ '\"]+", '', str(col)).strip()
    norm_map = {normalize(k): v for k, v in APPLY_HEADER_MAP.items()}
    new_cols = []
    for col in df.columns:
        ncol = normalize(col)
        new_cols.append(norm_map.get(ncol, col))
    df.columns = new_cols
    return df

def filter_apply_employees(df, emp_list):
    # Accept by name and/or ID when available
    if emp_list is None or emp_list.empty:
        return df
    emp_list = emp_list.copy()
    emp_list['Name'] = emp_list['Name'].astype(str)
    if 'ID Number' in emp_list.columns:
        emp_list['ID Number'] = emp_list['ID Number'].astype(str)

    valid_names = set(emp_list['Name'])
    valid_ids = set(emp_list['ID Number']) if 'ID Number' in emp_list.columns else set()

    # If neither Emp Name nor Emp ID present, nothing to filter
    if 'Emp Name' not in df.columns and 'Emp ID' not in df.columns:
        return df

    name_match = pd.Series([False] * len(df))
    id_match = pd.Series([False] * len(df))

    if 'Emp Name' in df.columns:
        name_match = df['Emp Name'].astype(str).isin(valid_names)
    if 'Emp ID' in df.columns and len(valid_ids) > 0:
        id_match = df['Emp ID'].astype(str).isin(valid_ids)

    return df[name_match | id_match]

def add_apply_columns(df):
    # Type
    if 'Application Type' in df.columns:
        def extract_type(val):
            if not isinstance(val, str):
                return ''
            val_lower = val.lower()
            # English/Chinese mapping
            if 'trip' in val_lower or 'trips' in val_lower or '出差' in val_lower:
                return 'Trip'
            if 'leave' in val_lower or '事假' in val_lower or '年休假' in val_lower:
                return 'Leave'
            if 'supp' in val_lower or 'forgot' in val_lower or 'forget' in val_lower or '补单' in val_lower:
                return 'Supp'
            if 'replenishment' in val_lower or '个人补单' in val_lower:
                return 'Replenishment'
            if '病假' in val_lower:
                return 'Sick leave'
            return ''
        df['Type'] = df['Application Type'].apply(extract_type)
    else:
        df['Type'] = ''
    # Bổ sung: Nếu Type vẫn rỗng, kiểm tra Note
    if 'Note' in df.columns:
        def fill_type_from_note(row):
            if row['Type']:
                return row['Type']
            note = str(row['Note']).lower()
            if any(x in note for x in ['supp', 'forgot', 'forget', '补单']):
                return 'Supp'
            if any(x in note for x in ['trip', 'trips', '出差']):
                return 'Trip'
            if any(x in note for x in ['leave', '事假', '年休假']):
                return 'Leave'
            if '病假' in note:
                return 'Sick leave'
            return row['Type']
        df['Type'] = df.apply(fill_type_from_note, axis=1)
    # Results
    if 'Approve Result' in df.columns:
        def map_approve_result(x):
            if pd.notna(x):
                x_str = str(x)
                if '通过' in x_str:
                    return 'Approved'
                elif '待审批' in x_str:
                    return 'Pending'
                elif '已撤销' in x_str:
                    return 'Withdrawal'
            return x
        df['Results'] = df['Approve Result'].apply(map_approve_result)
    else:
        df['Results'] = ''
    # Leave Type
    def map_leave_type_applytype(val):
        if pd.isna(val):
            return ''
        val_str = str(val)
        val_lower = val_str.lower()
        # English/Chinese mapping
        if 'sick' in val_lower or '病假' in val_lower:
            return 'Sick leave'
        if 'welfare' in val_lower:
            return 'Welfare'
        if 'annual' in val_lower or '年休假' in val_lower:
            return 'Annual leave'
        if 'leave' in val_lower or 'unpaid' in val_lower or '事假' in val_lower:
            return 'Leave'
        if any(x in val_str for x in ['市内出差', '国内出差', '国际/中国港澳台出差']):
            return 'Trip'
        if '补单' in val_lower or 'replenishment' in val_lower:
            return 'Replenishment'
        return val_str
    if 'Application Type' in df.columns:
        df['Leave Type'] = df['Application Type'].apply(map_leave_type_applytype)
    elif 'Note' in df.columns:
        df['Leave Type'] = df['Note'].apply(map_leave_type)
    else:
        df['Leave Type'] = ''
    return df

# =======================
# SUPPORT APPLY TRANSFORM
# =======================
def transform_support_apply_df(df, employee_list_df):
    """Transform Support team's apply file into standard apply format."""
    try:
        df = df.copy()

        # Preserve original status (might contain approval info)
        original_status_series = df['Status'] if 'Status' in df.columns else None

        # Normalize/rename straightforward columns
        rename_map = {}
        if 'Status' in df.columns:
            rename_map['Status'] = 'Application Type'
        if 'Reason' in df.columns:
            rename_map['Reason'] = 'Note'
        df = df.rename(columns=rename_map)

        # Helpers
        def parse_time_from_text(text):
            if pd.isna(text):
                return '00:00'
            s = str(text).strip()
            # Common phrases
            if '12am' in s.lower():
                return '00:00'
            if '12pm' in s.lower():
                return '12:00'
            # AM/PM like 9:30 AM or 9 AM
            m = re.search(r'(0?[1-9]|1[0-2])\s*[:.]\s*([0-5][0-9])\s*(AM|PM)', s, re.I)
            if m:
                hour = int(m.group(1))
                minute = int(m.group(2))
                ampm = m.group(3).upper()
                if ampm == 'PM' and hour != 12:
                    hour += 12
                if ampm == 'AM' and hour == 12:
                    hour = 0
                return f"{hour:02d}:{minute:02d}"
            m = re.search(r'(0?[1-9]|1[0-2])\s*(AM|PM)', s, re.I)
            if m:
                hour = int(m.group(1))
                ampm = m.group(2).upper()
                if ampm == 'PM' and hour != 12:
                    hour += 12
                if ampm == 'AM' and hour == 12:
                    hour = 0
                return f"{hour:02d}:00"
            # 24h format 14:30 or 9:05
            m = re.search(r'\b([01]?\d|2[0-3])\s*[:.]\s*([0-5]\d)\b', s)
            if m:
                hour = int(m.group(1))
                minute = int(m.group(2))
                return f"{hour:02d}:{minute:02d}"
            return '00:00'

        def extract_name_id_from_title(title_val):
            name_guess, emp_id_guess = None, None
            if pd.isna(title_val):
                return name_guess, emp_id_guess
            t = str(title_val)
            # Pattern like _@Name、12345678、
            m = re.search(r'@\s*([^、,]+)[、,]\s*(\d{7,12})', t)
            if m:
                return m.group(1).strip(), m.group(2)
            # Any 7-12 digit id
            m_id = re.search(r'(\d{7,12})', t)
            if m_id:
                emp_id_guess = m_id.group(1)
            # Try to take name between last '_' and the delimiter
            m_name = re.search(r'_(?:@)?\s*([^、,]+)', t)
            if m_name:
                name_guess = m_name.group(1).strip()
            return name_guess, emp_id_guess

        # Build normalized view of employee list for matching
        emp_df = employee_list_df.copy() if employee_list_df is not None else pd.DataFrame(columns=['Name','ID Number'])
        if not emp_df.empty:
            emp_df['NameStripped'] = emp_df['Name'].astype(str).str.replace(r'\d{7,}', '', regex=True).str.strip().str.lower()
            emp_df['ID Number'] = emp_df['ID Number'].astype(str)

        emp_names, emp_ids = [], []
        apply_dates, start_dates, end_dates = [], [], []

        for _, row in df.iterrows():
            # Title → Emp ID, Emp Name
            name_guess, id_guess = extract_name_id_from_title(row.get('Title', ''))
            chosen_name, chosen_id = '', ''
            if not emp_df.empty:
                match_row = None
                if id_guess:
                    # Try exact ID, then endswith
                    m = emp_df[emp_df['ID Number'] == id_guess]
                    if m.empty:
                        m = emp_df[emp_df['ID Number'].astype(str).str.endswith(str(id_guess))]
                    if not m.empty:
                        match_row = m.iloc[0]
                if match_row is None and name_guess:
                    norm_name = str(name_guess).replace('_',' ').strip().lower()
                    # Try stripped-name match
                    m = emp_df[emp_df['NameStripped'] == norm_name]
                    if m.empty:
                        m = emp_df[emp_df['Name'].astype(str).str.contains(re.escape(name_guess), case=False, na=False)]
                    if not m.empty:
                        match_row = m.iloc[0]
                if match_row is not None:
                    chosen_name = match_row['Name']
                    chosen_id = match_row.get('ID Number', '')
                else:
                    chosen_name = name_guess or ''
                    chosen_id = id_guess or ''
            else:
                chosen_name = name_guess or ''
                chosen_id = id_guess or ''
            emp_names.append(chosen_name)
            emp_ids.append(chosen_id)

            # Abnormal Day → Apply Date
            abnormal_day = row.get('Abnormal Day', '')
            try:
                day_dt = pd.to_datetime(abnormal_day)
                day_str = day_dt.strftime('%Y-%m-%d')
            except Exception:
                day_str = str(abnormal_day)
            apply_dates.append(day_str)

            # Start Date: Abnormal Day + time from From Note
            from_note = row.get('From Note', '')
            start_time = parse_time_from_text(from_note)
            start_dates.append(f"{day_str} {start_time}".strip())

            # End Date: Abnormal Day + time from To
            to_val = row.get('To', '')
            end_time = parse_time_from_text(to_val)
            end_dates.append(f"{day_str} {end_time}".strip())

        df['Emp ID'] = emp_ids
        df['Emp Name'] = emp_names
        if 'Name' not in df.columns:
            df['Name'] = df['Emp Name']
        df['Apply Date'] = apply_dates
        df['Start Date'] = start_dates
        df['End Date'] = end_dates

        # Define standard columns that should be kept
        std_cols = [
            'Emp ID', 'Emp Name', 'Submit ID', 'Submit Name', 
            'Apply Date', 'Start Date', 'End Date', 'Application Type', 
            'Note', 'Approve Result', 'Type', 'Results', 'Leave Type'
        ]
        
        # Initialize missing columns with empty values
        for col in std_cols:
            if col not in df.columns:
                df[col] = ''
        # Keep only the standard columns, drop all others
        df = df[std_cols]

        df = df[~df[['Application Type','Note','Approve Result']]
        .apply(lambda col: col.astype(str).str.strip().replace("nan","").isin(["","Please choose"]))
        .all(axis=1)]

        return df
    except Exception as e:
        print(f"Error transforming support apply file: {e}")
        return df

# =======================
# OT & LIEU
# =======================
def process_ot_lieu_df(df, employee_list_df):

    if 'Title' in df.columns:
        def extract_name_id(title):
            match = re.search(r'_([a-zA-Z\s]+?)[_\s](\d{7,8})', str(title))
            if match:
                emp_id = match.group(2)
                match_row = employee_list_df[employee_list_df['Name'].astype(str).str.endswith(emp_id)]
                if not match_row.empty:
                    return match_row.iloc[0]['Name']
            return None
        df['Name'] = df['Title'].apply(extract_name_id)
        cols = list(df.columns)
        if 'Name' in cols:
            cols.insert(0, cols.pop(cols.index('Name')))
            df = df[cols]

    ot_from_cols = [c for c in df.columns if re.search(r'ot.*from', c, re.I)]
    ot_to_cols = [c for c in df.columns if re.search(r'ot.*to', c, re.I)]
    sum_ot_col = next((c for c in df.columns if re.search(r'ot.*sum', c, re.I)), None)
    lieu_from_cols = [c for c in df.columns if re.search(r'lieu.*from', c, re.I)]
    lieu_to_cols = [c for c in df.columns if re.search(r'lieu.*to', c, re.I)]
    sum_lieu_col = next((c for c in df.columns if re.search(r'lieu.*sum', c, re.I)), None)
    
    def clean_hours(val):
        if pd.isna(val): return ''
        s = str(val).strip().lower()
        s = s.replace('hours', '').replace('hour', '').replace(',', '.').replace(';', '.').replace('；', '.').replace('：', ':')
        s = s.replace('h', ':')
        s = re.sub(r'[^0-9.:]', '', s)

        if ':' in s:
            parts = s.split(':')
            try:
                hour = int(parts[0])
                minute = int(parts[1]) if len(parts) > 1 else 0
                return round(hour + minute/60, 2)
            except:
                return s
        try:
            return float(s)
        except:
            return s
    if sum_ot_col:
        df[sum_ot_col] = df[sum_ot_col].apply(clean_hours)
    if sum_lieu_col:
        df[sum_lieu_col] = df[sum_lieu_col].apply(clean_hours)

    # Mark warnings for invalid time format
    def mark_cell(val, error=False, suggest=None, warning=False):
        if error:
            return {'value': val, 'error': True, 'suggest': suggest}
        if warning:
            return {'value': val, 'warning': True}
        return val

    # Helper function to parse time formats (consolidated)
    def parse_time_to_24h(val):
        """Parse various time formats to HH:MM 24-hour format"""
        if pd.isna(val) or not str(val).strip():
            return ''
        s = str(val).strip()
        
        # AM/PM format: 9:30 AM, 2:45 PM
        m = re.match(r'^(0?[1-9]|1[0-2])\s*[:. ]\s*([0-5][0-9])\s*(AM|PM|am|pm)$', s)
        if m:
            hour = int(m.group(1))
            minute = int(m.group(2))
            ampm = m.group(3).upper()
            if ampm == 'PM' and hour != 12:
                hour += 12
            if ampm == 'AM' and hour == 12:
                hour = 0
            return f'{hour:02d}:{minute:02d}'
        
        # 24-hour format: 14:30, 09:15
        m = re.match(r'^([01]?[0-9]|2[0-3])\s*[:. h]\s*([0-5][0-9])$', s)
        if m:
            hour = int(m.group(1))
            minute = int(m.group(2))
            return f'{hour:02d}:{minute:02d}'
        
        return None

    # Normalize time format for OT/Lieu columns
    def norm_time_to_24h(val):
        if str(val).strip() in ['Hour : Minutes AM', 'Hour : Minutes PM', 'Hour h Min']:
            return val
        parsed = parse_time_to_24h(val)
        if parsed is not None and parsed != '':
            return parsed
        elif val and str(val).strip():
            return mark_cell(val, warning=True)
        else:
            return val

    # Apply time normalization
    for col in ot_from_cols + ot_to_cols + lieu_from_cols + lieu_to_cols:
        df[col] = df[col].apply(norm_time_to_24h)


    # Helper function to calculate time difference with lunch break deduction
    def calculate_time_diff(time_from, time_to):
        """Calculate time difference in hours with lunch break deduction"""
        if re.match(r'^([01]?\d|2[0-3]):[0-5]\d$', str(time_from)) and re.match(r'^([01]?\d|2[0-3]):[0-5]\d$', str(time_to)):
            t1 = datetime.strptime(str(time_from), '%H:%M')
            t2 = datetime.strptime(str(time_to), '%H:%M')
            diff = (t2 - t1).total_seconds() / 3600
            if diff < 0:
                diff += 24
            # Deduct 1.5h lunch break if spans lunch time
            if t1.hour <= 12 < t2.hour or (t1.hour == 12 and t2.hour > 13):
                if t2.hour > 13 or (t2.hour == 13 and t2.minute >= 30):
                    diff -= 1.5
            return round(diff, 2)
        return None

    # Process OT time calculation (consolidated)
    if ot_from_cols and ot_to_cols and sum_ot_col:
        for idx, row in df.iterrows():
            ot_from = row[ot_from_cols[0]] if not isinstance(row[ot_from_cols[0]], dict) else row[ot_from_cols[0]].get('value')
            ot_to = row[ot_to_cols[0]] if not isinstance(row[ot_to_cols[0]], dict) else row[ot_to_cols[0]].get('value')
            user_val = row[sum_ot_col] if not isinstance(row[sum_ot_col], dict) else row[sum_ot_col].get('value')
            
            real_hours = calculate_time_diff(ot_from, ot_to)
            if real_hours is not None:
                try:
                    user_val_f = float(user_val)
                except:
                    user_val_f = None
                if user_val_f is None or abs(real_hours - user_val_f) > 0.01:
                    df.at[idx, sum_ot_col] = mark_cell(user_val, error=True, suggest=real_hours)
                else:
                    df.at[idx, sum_ot_col] = real_hours

    # Process Lieu time calculation (consolidated)
    if lieu_from_cols and lieu_to_cols and sum_lieu_col:
        for idx, row in df.iterrows():
            lieu_from = row[lieu_from_cols[0]] if not isinstance(row[lieu_from_cols[0]], dict) else row[lieu_from_cols[0]].get('value')
            lieu_to = row[lieu_to_cols[0]] if not isinstance(row[lieu_to_cols[0]], dict) else row[lieu_to_cols[0]].get('value')
            user_val = row[sum_lieu_col] if not isinstance(row[sum_lieu_col], dict) else row[sum_lieu_col].get('value')
            
            real_hours = calculate_time_diff(lieu_from, lieu_to)
            if real_hours is not None:
                try:
                    user_val_f = float(user_val)
                except:
                    user_val_f = None
                if user_val_f is None or abs(real_hours - user_val_f) > 0.01:
                    df.at[idx, sum_lieu_col] = mark_cell(user_val, error=True, suggest=real_hours)
                else:
                    df.at[idx, sum_lieu_col] = real_hours

    special_vals = ['Hour : Minutes AM', 'Hour : Minutes PM', 'Hour h Min']
    def is_empty_or_special(val):
        return pd.isna(val) or str(val).strip() == '' or str(val).strip() in special_vals
    if ot_from_cols and ot_to_cols and lieu_from_cols and lieu_to_cols:
        df = df[~(
            df[ot_from_cols[0]].apply(is_empty_or_special) &
            df[ot_to_cols[0]].apply(is_empty_or_special) &
            df[lieu_from_cols[0]].apply(is_empty_or_special) &
            df[lieu_to_cols[0]].apply(is_empty_or_special)
        )].reset_index(drop=True)

    def mark_gray(val):
        return {'value': val, 'gray': True}
    
    ot_day_col = next((c for c in df.columns if re.search(r'ot.*day', c, re.I)), None)
    lieu_date_col = next((c for c in df.columns if re.search(r'lieu.*date', c, re.I)), None)

    # OT From/To
    if ot_from_cols and ot_to_cols:
        for idx, row in df.iterrows():
            ot_from = row[ot_from_cols[0]]
            ot_to = row[ot_to_cols[0]]
            if is_empty_or_special(ot_from) and is_empty_or_special(ot_to):
                df.at[idx, ot_from_cols[0]] = mark_gray(ot_from)
                df.at[idx, ot_to_cols[0]] = mark_gray(ot_to)
                if sum_ot_col:
                    df.at[idx, sum_ot_col] = mark_gray(row[sum_ot_col])
                if ot_day_col:
                    df.at[idx, ot_day_col] = mark_gray(row[ot_day_col])
    # Lieu From/To
    if lieu_from_cols and lieu_to_cols:
        for idx, row in df.iterrows():
            lieu_from = row[lieu_from_cols[0]]
            lieu_to = row[lieu_to_cols[0]]
            if is_empty_or_special(lieu_from) and is_empty_or_special(lieu_to):
                df.at[idx, lieu_from_cols[0]] = mark_gray(lieu_from)
                df.at[idx, lieu_to_cols[0]] = mark_gray(lieu_to)
                if sum_lieu_col:
                    df.at[idx, sum_lieu_col] = mark_gray(row[sum_lieu_col])
                if lieu_date_col:
                    df.at[idx, lieu_date_col] = mark_gray(row[lieu_date_col])
    return df

# ========================
# UPLOAD & SAVE EXCEL
# ========================
def load_excel_data(file_path, sheet_name=None):
    """Load data from Excel file with support for .xls and .xlsx"""
    try:
        file_ext = os.path.splitext(file_path)[1].lower()
        if file_ext == '.xls':
            try:
                if sheet_name:
                    df = pd.read_excel(file_path, sheet_name=sheet_name, engine='xlrd')
                else:
                    df = pd.read_excel(file_path, engine='xlrd')
            except Exception as e:
                print(f"xlrd engine failed for .xls file: {e}")
                try:
                    if sheet_name:
                        df = pd.read_excel(file_path, sheet_name=sheet_name, engine='openpyxl')
                    else:
                        df = pd.read_excel(file_path, engine='openpyxl')
                except Exception as e2:
                    print(f"openpyxl engine also failed for .xls file: {e2}")
                    raise Exception(f"Cannot read .xls file. Please convert to .xlsx format. Error: {str(e)}")
        else:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            
            if sheet_name:
                if sheet_name in workbook.sheetnames:
                    df = pd.read_excel(file_path, sheet_name=sheet_name)
                else:
                    return None
            else:
                df = pd.read_excel(file_path, sheet_name=workbook.sheetnames[0])
        
        # Clean the data
        df = df.dropna(how='all')
        df = df.fillna('')
        return df
    except Exception as e:
        print(f"Error loading Excel file: {e}")
        return None

# ========================
# APP ROUTE
# ========================
@app.route('/')
def index():
    """Main page"""
    return render_template('index.html')

@app.route('/import/signinout', methods=['POST'])
def import_signinout():
    """Import Sign In/Out data (only emp_name and attendance_time columns)"""
    global sign_in_out_data
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    if file and file.filename.lower().endswith(('.xlsx', '.xls', '.csv')):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        if file.filename.lower().endswith('.csv'):
            df = try_read_csv(open(file_path, 'rb').read())
        elif file.filename.lower().endswith('.xls'):
            try:
                df = pd.read_excel(file_path, engine='xlrd')
            except Exception as e:
                print(f"xlrd engine failed: {e}")
                try:
                    df = pd.read_excel(file_path, engine='openpyxl')
                except Exception as e2:
                    print(f"openpyxl engine also failed: {e2}")
                    return jsonify({'error': f'Cannot read .xls file. Please convert to .xlsx format. Error: {str(e)}'}), 400
        else:
            df = pd.read_excel(file_path)

        keep = [col for col in df.columns if col.lower() in ['emp_name', 'attendance_time']]
        df = df[keep]
        sign_in_out_data = df.reset_index(drop=True)
        return jsonify({
            'success': True,
            'message': f'Sign In/Out data imported successfully. Loaded {len(df)} rows.',
            'rows': len(df)
        })
    else:
        return jsonify({'error': 'Invalid file type'}), 400

@app.route('/import/apply', methods=['POST'])
def import_apply():
    """Import Apply data with header translation, employee filter, and extra columns"""
    global apply_data, employee_list_df
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    if file and file.filename.lower().endswith(('.xlsx', '.xls', '.csv')):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        if file.filename.lower().endswith('.csv'):
            df = try_read_csv(open(file_path, 'rb').read())
        elif file.filename.lower().endswith('.xls'):
            try:
                df = pd.read_excel(file_path, engine='xlrd')
            except Exception as e:
                print(f"xlrd engine failed: {e}")
                try:
                    df = pd.read_excel(file_path, engine='openpyxl')
                except Exception as e2:
                    print(f"openpyxl engine also failed: {e2}")
                    return jsonify({'error': f'Cannot read .xls file. Please convert to .xlsx format. Error: {str(e)}'}), 400
        else:
            df = pd.read_excel(file_path)
        # Transform for support files
        if 'support' in filename.lower() or 'suport' in filename.lower():
            df = transform_support_apply_df(df, employee_list_df)
        df = translate_apply_headers(df)
        df = filter_apply_employees(df, employee_list_df)
        df = add_apply_columns(df)
        apply_data = df.reset_index(drop=True)
        return jsonify({
            'success': True,
            'message': f'Apply data imported successfully. Loaded {len(df)} rows.',
            'rows': len(df)
        })
    else:
        return jsonify({'error': 'Invalid file type'}), 400

@app.route('/import/otlieu', methods=['POST'])
def import_otlieu():
    global ot_lieu_data, employee_list_df
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    if file and file.filename.lower().endswith(('.xlsx', '.xls', '.csv')):
        filename = secure_filename(file.filename)
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(file_path)
        data = load_excel_data(file_path)
        if data is not None:
            if 'OT From Note: 12AM is midnight' in data.columns:
                data = data.rename(columns={'OT From Note: 12AM is midnight': 'OT From'})
            data = process_ot_lieu_df(data, employee_list_df)
            ot_lieu_data = data
            return jsonify({
                'success': True,
                'message': f'OT Lieu data imported successfully. Loaded {len(data)} rows.',
                'rows': len(data)
            })
        else:
            return jsonify({'error': 'Failed to load file'}), 400
    else:
        return jsonify({'error': 'Invalid file type'}), 400
    
@app.route('/clear/signinout', methods=['POST'])
def clear_signinout():
    """Clear Sign In/Out data"""
    global sign_in_out_data
    sign_in_out_data = None
    return jsonify({'success': True, 'message': 'Sign In/Out data cleared successfully'})

@app.route('/clear/apply', methods=['POST'])
def clear_apply():
    """Clear Apply data"""
    global apply_data
    apply_data = None
    return jsonify({'success': True, 'message': 'Apply data cleared successfully'})

@app.route('/clear/otlieu', methods=['POST'])
def clear_otlieu():
    """Clear OT Lieu data"""
    global ot_lieu_data
    ot_lieu_data = None
    return jsonify({'success': True, 'message': 'OT Lieu data cleared successfully'})

@app.route('/export', methods=['GET'])
def export():
    """
    Export processed data as Excel file using existing template.
    Write data to existing sheets, keep all formatting/colors.
    """
    global sign_in_out_data, apply_data, ot_lieu_data, employee_list_df, rules

    try:
        # Get month/year parameters from request
        selected_month = request.args.get('month', type=int)
        selected_year = request.args.get('year', type=int)
        
        # Đường dẫn template gốc
        template_path = os.path.join(app.config['UPLOAD_FOLDER'], 'AttendanceReport.xlsx')
        if not os.path.exists(template_path):
            return jsonify({'error': 'Template file AttendanceReport.xlsx not found'}), 400

        # Tạo file mới từ template
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"AttendanceReport_{timestamp}.xlsx"
        file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        import shutil
        shutil.copy2(template_path, file_path)

        # Load workbook với openpyxl (giữ định dạng)
        from openpyxl import load_workbook
        wb = load_workbook(file_path)

        # Helper: Ghi DataFrame vào worksheet, giữ header, chỉ ghi dữ liệu từ dòng 2
        def write_df_to_sheet(ws, df, start_row=2):
            # Xóa dữ liệu cũ (giữ header)
            for row in ws.iter_rows(min_row=start_row, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.value = None
            # Ghi dữ liệu mới
            for idx, row in df.iterrows():
                for col_idx, value in enumerate(row, 1):
                    ws.cell(row=idx + start_row, column=col_idx, value=value)

        # 1. Employee List
        if employee_list_df is not None and not employee_list_df.empty and 'Emp List' in wb.sheetnames:
            write_df_to_sheet(wb['Emp List'], employee_list_df, start_row=2)

        # 2. Rules
        if rules is not None and not rules.empty and 'Rules' in wb.sheetnames:
            write_df_to_sheet(wb['Rules'], rules, start_row=2)

        # 3. Sign In-Out Data
        if sign_in_out_data is not None and not sign_in_out_data.empty and 'Sign in-out data' in wb.sheetnames:
            write_df_to_sheet(wb['Sign in-out data'], sign_in_out_data, start_row=2)

        # 4. Apply Data
        if apply_data is not None and not apply_data.empty and 'Apply data' in wb.sheetnames:
            write_df_to_sheet(wb['Apply data'], apply_data, start_row=2)

        # 6. OT Lieu Before (calculated)
        if 'OT Lieu data' in wb.sheetnames:
            try:
                otlieu_before_df = calculate_otlieu_before()
                if otlieu_before_df is not None and not otlieu_before_df.empty:
                    write_df_to_sheet(wb['OT Lieu data'], otlieu_before_df, start_row=3)
            except Exception as e:
                print(f"Error calculating OT Lieu Before: {e}")

        # 7. OT Lieu Report (calculated)
        if 'OT & Lieu Report' in wb.sheetnames:
            try:
                otlieu_report_result = calculate_otlieu_report_for_export()
                if isinstance(otlieu_report_result, dict) and 'columns' in otlieu_report_result and 'rows' in otlieu_report_result:
                    otlieu_report_df = pd.DataFrame(otlieu_report_result['rows'], columns=otlieu_report_result['columns'])
                    if not otlieu_report_df.empty:
                        write_df_to_sheet(wb['OT & Lieu Report'], otlieu_report_df, start_row=9)
            except Exception as e:
                print(f"Error calculating OT Lieu Report: {e}")

        # 8. Total Attendance Detail (calculated)
        if 'Total Attendance detail' in wb.sheetnames:
            try:
                total_attendance_result = calculate_total_attendance_detail_for_export(selected_month, selected_year)
                if isinstance(total_attendance_result, dict) and 'columns' in total_attendance_result and 'rows' in total_attendance_result:
                    total_attendance_df = pd.DataFrame(total_attendance_result['rows'], columns=total_attendance_result['columns'])
                    if not total_attendance_df.empty:
                        write_df_to_sheet(wb['Total Attendance detail'], total_attendance_df, start_row=5)
            except Exception as e:
                print(f"Error calculating Total Attendance Detail: {e}")

        # 9. Abnormal Late/Early Data
        if 'Abnormal LateCome-EarlyLeave' in wb.sheetnames:
            try:
                abnormal_late_early_result = calculate_abnormal_late_early_for_export(selected_month, selected_year)
                if isinstance(abnormal_late_early_result, dict) and 'columns' in abnormal_late_early_result and 'rows' in abnormal_late_early_result:
                    abnormal_late_early_df = pd.DataFrame(abnormal_late_early_result['rows'], columns=abnormal_late_early_result['columns'])
                    if not abnormal_late_early_df.empty:
                        write_df_to_sheet(wb['Abnormal LateCome-EarlyLeave'], abnormal_late_early_df, start_row=2)
            except Exception as e:
                print(f"Error calculating Abnormal Late/Early data: {e}")

        # 10. Abnormal Missing Data
        if 'Abnormal Missing' in wb.sheetnames:
            try:
                abnormal_missing_df = calculate_abnormal_missing_for_export(selected_month, selected_year)
                if abnormal_missing_df is not None and not abnormal_missing_df.empty:
                    write_df_to_sheet(wb['Abnormal Missing'], abnormal_missing_df, start_row=2)
            except Exception as e:
                print(f"Error calculating Abnormal Missing data: {e}")

        # Lưu file và trả về
        wb.save(file_path)
        wb.close()
        return send_file(file_path, as_attachment=True, download_name=filename)

    except Exception as e:
        return jsonify({'error': f'Failed to export file: {str(e)}'}), 400

def try_read_csv(file_bytes, **kwargs):
    encodings = ['utf-8', 'utf-8-sig', 'cp1252', 'latin1', 'gbk']
    for enc in encodings:
        try:
            return pd.read_csv(BytesIO(file_bytes), encoding=enc, **kwargs)
        except Exception:
            continue
    raise ValueError('Could not decode CSV file with common encodings.')

@app.route('/preview_upload', methods=['POST'])
def preview_upload():
    """Preview uploaded file: return sheet names and preview data for each sheet"""
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    filename = secure_filename(file.filename)
    ext = os.path.splitext(filename)[1].lower()
    file_bytes = file.read()
    preview = {}
    sheet_names = []
    try:
        if ext in ['.xlsx', '.xls']:
            if ext == '.xls':
                try:
                    excel = pd.ExcelFile(BytesIO(file_bytes), engine='xlrd')
                except Exception as e:
                    print(f"xlrd engine failed for .xls preview: {e}")
                    try:
                        excel = pd.ExcelFile(BytesIO(file_bytes), engine='openpyxl')
                    except Exception as e2:
                        print(f"openpyxl engine also failed for .xls preview: {e2}")
                        return jsonify({'error': f'Cannot read .xls file. Please convert to .xlsx format. Error: {str(e)}'}), 400
            else:
                excel = pd.ExcelFile(BytesIO(file_bytes))
            sheet_names = excel.sheet_names
            for sheet in sheet_names:
                df = pd.read_excel(excel, sheet_name=sheet, nrows=10)
                preview[sheet] = {
                    'columns': df.columns.tolist(),
                    'rows': df.fillna('').astype(str).values.tolist()
                }
        elif ext == '.csv':
            df = try_read_csv(file_bytes, nrows=10)
            preview['CSV'] = {
                'columns': df.columns.tolist(),
                'rows': df.fillna('').astype(str).values.tolist()
            }
            sheet_names = ['CSV']
        else:
            return jsonify({'error': 'Unsupported file type'}), 400
        return jsonify({'success': True, 'sheet_names': sheet_names, 'preview': preview})
    except Exception as e:
        return jsonify({'error': f'Failed to preview file: {str(e)}'}), 400

@app.route('/import_with_sheet', methods=['POST'])
def import_with_sheet():
    """Import a specific sheet from an uploaded file for a given data type"""
    global sign_in_out_data, apply_data, ot_lieu_data, employee_list_df
    data_type = request.form.get('data_type')
    sheet_name = request.form.get('sheet_name')
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    filename = secure_filename(file.filename)
    ext = os.path.splitext(filename)[1].lower()
    file_bytes = file.read()
    try:
        if ext in ['.xlsx', '.xls']:
            if ext == '.xls':
                try:
                    df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name, engine='xlrd')
                except Exception as e:
                    print(f"xlrd engine failed for .xls import: {e}")
                    try:
                        df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name, engine='openpyxl')
                    except Exception as e2:
                        print(f"openpyxl engine also failed for .xls import: {e2}")
                        return jsonify({'error': f'Cannot read .xls file. Please convert to .xlsx format. Error: {str(e)}'}), 400
            else:
                df = pd.read_excel(BytesIO(file_bytes), sheet_name=sheet_name)
        elif ext == '.csv':
            df = try_read_csv(file_bytes)
        else:
            return jsonify({'error': 'Unsupported file type'}), 400
        df = df.dropna(how='all').fillna('')
        if data_type == 'signinout':
            sign_in_out_data = df
        elif data_type == 'apply':
            # Transform for support files (use filename)
            if 'support' in filename.lower() or 'suport' in filename.lower():
                df = transform_support_apply_df(df, employee_list_df)
            df = translate_apply_headers(df)
            df = filter_apply_employees(df, employee_list_df)
            df = add_apply_columns(df)
            apply_data = df.reset_index(drop=True)
        elif data_type == 'otlieu':
            if 'OT From Note: 12AM is midnight' in df.columns:
                df = df.rename(columns={'OT From Note: 12AM is midnight': 'OT From'})
            df = process_ot_lieu_df(df, employee_list_df)
            ot_lieu_data = df
        else:
            return jsonify({'error': 'Invalid data type'}), 400
        return jsonify({'success': True, 'message': f'{data_type.capitalize()} data imported successfully. Loaded {len(df)} rows.', 'rows': len(df)})
    except Exception as e:
        print(str(e))
        return jsonify({'error': f'Failed to import file: {str(e)}'}), 400

@app.route('/employee_list', methods=['GET'])
def employee_list():
    global employee_list_df
    df = employee_list_df.copy()
    if df.empty:
        return Markup('<div class="text-muted">No employees loaded.</div>')
    html = '<table class="table table-bordered table-sm employee-table"><thead><tr>'
    for col in df.columns:
        html += f'<th>{col}</th>'
    html += '</tr></thead><tbody>'
    for idx, row in df.iterrows():
        html += '<tr>'
        for col in df.columns:
            html += f'<td>{row[col]}</td>'
        html += f'<td><button class="hover-delete-btn" onclick="removeEmployee({idx})" title="Delete employee">×</button></td>'
        html += '</tr>'
    html += '</tbody></table>'
    return Markup(html)

@app.route('/employee_list_json', methods=['GET'])
def employee_list_json():
    """Return employee list as JSON for department filter"""
    global employee_list_df
    if employee_list_df is None or employee_list_df.empty:
        return jsonify({'data': []})
    
    # Convert to list of dictionaries
    data = []
    for _, row in employee_list_df.iterrows():
        data.append({
            'Name': str(row.get('Name', '')),
            'ID Number': str(row.get('ID Number', '')),
            'Dept': str(row.get('Dept', '')),
            'Internship': str(row.get('Internship', ''))
        })
    
    return jsonify({'data': data})

@app.route('/employee_list_filtered', methods=['GET'])
def employee_list_filtered():
    global employee_list_df
    df = employee_list_df.copy()
    
    # Get filter parameters
    dept_filter = request.args.get('dept', '')
    intern_filter = request.args.get('intern', '')
    
    # Apply filters
    if dept_filter:
        df = df[df['Dept'].astype(str).str.contains(dept_filter, case=False, na=False)]
    
    if intern_filter:
        if intern_filter == 'Intern':
            df = df[df['Internship'].astype(str).str.contains('Intern', case=False, na=False)]
        elif intern_filter == 'Regular':
            df = df[~df['Internship'].astype(str).str.contains('Intern', case=False, na=False)]
    
    if df.empty:
        return Markup('<div class="text-muted">No employees match the filter criteria.</div>')
    
    html = '<table class="table table-bordered table-sm employee-table"><thead><tr>'
    for col in df.columns:
        html += f'<th>{col}</th>'
    html += '</tr></thead><tbody>'
    
    for idx, row in df.iterrows():
        html += '<tr>'
        for col in df.columns:
            html += f'<td>{row[col]}</td>'
        html += f'<td><button class="hover-delete-btn" onclick="removeEmployee({idx})" title="Delete employee">×</button></td>'
        html += '</tr>'
    html += '</tbody></table>'
    
    # Add filter summary
    filter_summary = f'<div class="text-muted small mb-2">Showing {len(df)} of {len(employee_list_df)} employees'
    if dept_filter or intern_filter:
        filter_summary += ' (filtered)'
    filter_summary += '</div>'
    
    return Markup(filter_summary + html)

@app.route('/upload_employee_list', methods=['POST'])
def upload_employee_list():
    global employee_list_df
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    ext = os.path.splitext(file.filename)[1].lower()
    file_bytes = file.read()
    try:
        if ext in ['.xlsx', '.xls']:
            df = pd.read_excel(BytesIO(file_bytes), dtype=str)
        elif ext == '.csv':
            df = try_read_csv(file_bytes, dtype=str)
        else:
            return jsonify({'error': 'Unsupported file type'}), 400
        keep = ["STT", "Name", "ID Number", "Dept", "Internship"]
        for col in keep:
            if col not in df.columns:
                df[col] = ''
        df = df[keep]
        df = df.fillna('')
        if "ID Number" in df.columns:
            df["ID Number"] = df["ID Number"].astype(str)
        employee_list_df = df.reset_index(drop=True)
        # Save to file
        employee_list_df.to_csv(EMPLOYEE_LIST_PATH, index=False)
        return jsonify({'success': True, 'message': f'Imported {len(df)} employees.'})
    except Exception as e:
        return jsonify({'error': f'Failed to import employee list: {str(e)}'}), 400

@app.route('/add_employee', methods=['POST'])
def add_employee():
    global employee_list_df
    try:
        data = request.json
        columns = list(employee_list_df.columns)
        row = [''] * len(columns)
        for idx, col in enumerate(columns):
            if col == 'Name':
                row[idx] = data.get('Name', '')
            elif col == 'ID Number':
                row[idx] = str(data.get('ID Number', ''))
            elif col == 'Dept':
                row[idx] = data.get('Dept', '')
            elif col == 'Internship':
                row[idx] = data.get('Internship', '')
        employee_list_df.loc[len(employee_list_df)] = row
        if 'STT' in employee_list_df.columns:
            employee_list_df['STT'] = range(1, len(employee_list_df) + 1)
        employee_list_df.to_csv(EMPLOYEE_LIST_PATH, index=False)
        return jsonify({'success': True, 'message': 'Employee added.'})
    except Exception as e:
        return jsonify({'error': f'Failed to add employee: {str(e)}'}), 400

@app.route('/remove_employee', methods=['POST'])
def remove_employee():
    global employee_list_df
    try:
        idx = int(request.json.get('index'))
        employee_list_df = employee_list_df.drop(idx).reset_index(drop=True)
        # Auto fill STT ascending if exists
        if 'STT' in employee_list_df.columns:
            employee_list_df['STT'] = range(1, len(employee_list_df) + 1)
        employee_list_df.to_csv(EMPLOYEE_LIST_PATH, index=False)
        return jsonify({'success': True, 'message': 'Employee removed.'})
    except Exception as e:
        return jsonify({'error': f'Failed to remove employee: {str(e)}'}), 400

@app.route('/get_signinout_data')
def get_signinout_data():
    global sign_in_out_data
    if sign_in_out_data is None or sign_in_out_data.empty:
        return jsonify({'columns': [], 'data': []})
    cols = list(sign_in_out_data.columns)
    rows = sign_in_out_data.fillna('').astype(str).values.tolist()
    return jsonify({'columns': cols, 'data': rows})

@app.route('/get_apply_data')
def get_apply_data():
    global apply_data
    if apply_data is None or apply_data.empty:
        return jsonify({'columns': [], 'data': []})
    cols = list(apply_data.columns)
    rows = apply_data.fillna('').astype(str).values.tolist()
    return jsonify({'columns': cols, 'data': rows})

@app.route('/get_otlieu_data')
def get_otlieu_data():
    global ot_lieu_data
    if ot_lieu_data is None or ot_lieu_data.empty:
        return jsonify({'columns': [], 'data': []})
    cols = list(ot_lieu_data.columns)
    rows = ot_lieu_data.fillna('').values.tolist()
    return jsonify({'columns': cols, 'data': rows})

@app.route('/update_apply_row', methods=['POST'])
def update_apply_row():
    global apply_data
    try:
        data = request.json
        idx = int(data.get('index'))
        col = data.get('column')
        value = data.get('value')
        if apply_data is not None and 0 <= idx < len(apply_data) and col in apply_data.columns:
            apply_data.at[idx, col] = value
            return jsonify({'success': True})
        else:
            return jsonify({'error': 'Invalid index or column'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/update_otlieu_row', methods=['POST'])
def update_otlieu_row():
    global ot_lieu_data
    try:
        data = request.json
        idx = int(data.get('index'))
        col = data.get('column')
        value = data.get('value')
        if ot_lieu_data is not None and 0 <= idx < len(ot_lieu_data) and col in ot_lieu_data.columns:
            ot_lieu_data.at[idx, col] = value
            updated = {}
            ot_from_cols = [c for c in ot_lieu_data.columns if re.search(r'ot.*from', c, re.I)]
            ot_to_cols = [c for c in ot_lieu_data.columns if re.search(r'ot.*to', c, re.I)]
            sum_ot_col = next((c for c in ot_lieu_data.columns if re.search(r'ot.*sum', c, re.I)), None)
            lieu_from_cols = [c for c in ot_lieu_data.columns if re.search(r'lieu.*from', c, re.I)]
            lieu_to_cols = [c for c in ot_lieu_data.columns if re.search(r'lieu.*to', c, re.I)]
            sum_lieu_col = next((c for c in ot_lieu_data.columns if re.search(r'lieu.*sum', c, re.I)), None)
            # OT
            if col in ot_from_cols + ot_to_cols and sum_ot_col:
                ot_from = ot_lieu_data.at[idx, ot_from_cols[0]]
                ot_to = ot_lieu_data.at[idx, ot_to_cols[0]]
                if re.match(r'^([01]?\d|2[0-3]):[0-5]\d$', str(ot_from)) and re.match(r'^([01]?\d|2[0-3]):[0-5]\d$', str(ot_to)):
                    t1 = datetime.strptime(str(ot_from), '%H:%M')
                    t2 = datetime.strptime(str(ot_to), '%H:%M')
                    diff = (t2 - t1).total_seconds() / 3600
                    if diff < 0:
                        diff += 24
                    if t1.hour <= 12 < t2.hour or (t1.hour == 12 and t2.hour > 13):
                        if t2.hour > 13 or (t2.hour == 13 and t2.minute >= 30):
                            diff -= 1.5
                    real = round(diff, 2)
                    ot_lieu_data.at[idx, sum_ot_col] = real
                    updated[sum_ot_col] = real

            # Lieu
            if col in lieu_from_cols + lieu_to_cols and sum_lieu_col:
                lieu_from = ot_lieu_data.at[idx, lieu_from_cols[0]]
                lieu_to = ot_lieu_data.at[idx, lieu_to_cols[0]]
                if re.match(r'^([01]?\d|2[0-3]):[0-5]\d$', str(lieu_from)) and re.match(r'^([01]?\d|2[0-3]):[0-5]\d$', str(lieu_to)):
                    t1 = datetime.strptime(str(lieu_from), '%H:%M')
                    t2 = datetime.strptime(str(lieu_to), '%H:%M')
                    diff = (t2 - t1).total_seconds() / 3600
                    if diff < 0:
                        diff += 24
                    if t1.hour <= 12 < t2.hour or (t1.hour == 12 and t2.hour > 13):
                        if t2.hour > 13 or (t2.hour == 13 and t2.minute >= 30):
                            diff -= 1.5
                    real = round(diff, 2)
                    ot_lieu_data.at[idx, sum_lieu_col] = real
                    updated[sum_lieu_col] = real

            # thêm sau khi chỉnh sửa
            temp_path = os.path.join(app.config['UPLOAD_FOLDER'], 'temp_otlieu.xlsx')
            flat_df = ot_lieu_data.applymap(flatten_cell)
            flat_df.to_excel(temp_path, index=False)

            return jsonify({'success': True, 'updated': updated})
        else:
            return jsonify({'error': 'Invalid index or column'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/get_apply_column_options')
def get_apply_column_options():
    global apply_data
    default_type = ["Trip", "Leave", "Supp", "Replenishment", ""]
    default_leave_type = ["Unpaid", "Sick", "Welfare", "Annual", ""]
    leave_type_values = []
    if apply_data is not None and not apply_data.empty and 'Leave Type' in apply_data.columns:
        leave_type_values = [v for v in apply_data['Leave Type'].unique() if str(v).strip() != '']
    leave_type_all = default_leave_type + [v for v in leave_type_values if v not in default_leave_type]
    
    return jsonify({"Type": default_type, "Leave Type": leave_type_all})

def get_holidays_from_rules():
    global rules
    try:
        if rules is not None and 'Holiday Date in This Year' in rules.columns:
            holidays = []
            for val in rules['Holiday Date in This Year']:
                if pd.notna(val):
                    try:
                        d = pd.to_datetime(val)
                        holidays.append(d.strftime('%m-%d'))
                    except:
                        pass
            return set(holidays)
    except Exception as e:
        print("Error reading holidays from rules:", e)
    return set()

# ----------------------------
# ROUTE - RULES PAGE
# ----------------------------
@app.route('/get_rules_table')
def get_rules_table():
    global rules
    try:
        if rules is not None:
            df = rules.fillna('')
            columns = list(df.columns)
            rows = df.astype(str).values.tolist()
            return jsonify({'columns': columns, 'rows': rows})
        else:
            return jsonify({'columns': [], 'rows': [], 'error': 'No rules loaded'})
    except Exception as e:
        return jsonify({'columns': [], 'rows': [], 'error': str(e)})

@app.route('/update_rule_cell', methods=['POST'])
def update_rule_cell():
    global rules
    try:
        data = request.json
        row = int(data['row'])
        col = int(data['col'])
        value = data['value']
        if rules is not None:
            rules.iat[row, col] = value
            rules.to_excel(RULES_XLSX_PATH, index=False)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/delete_rule_row', methods=['POST'])
def delete_rule_row():
    global rules
    try:
        data = request.json
        row = int(data['row'])
        if rules is not None:
            rules = rules.drop(rules.index[row]).reset_index(drop=True)
            rules.to_excel(RULES_XLSX_PATH, index=False)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/add_rule_row', methods=['POST'])
def add_rule_row():
    global rules
    try:
        if rules is not None:
            new_row = ['' for _ in rules.columns]
            rules.loc[len(rules)] = new_row
            rules.to_excel(RULES_XLSX_PATH, index=False)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/batch_update_rule_cells', methods=['POST'])
def batch_update_rule_cells():
    global rules
    try:
        data = request.json
        edits = data.get('edits', [])
        if rules is not None:
            for edit in edits:
                row = int(edit['row'])
                col = int(edit['col'])
                value = edit['value']
                rules.iat[row, col] = value
            rules.to_excel(RULES_XLSX_PATH, index=False)
        return jsonify({'success': True})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/reload_rules', methods=['POST'])
def reload_rules():
    """Reload rules from file"""
    global rules
    try:
        if os.path.exists(RULES_XLSX_PATH):
            rules = pd.read_excel(RULES_XLSX_PATH)
            return jsonify({'success': True, 'message': 'Rules reloaded successfully'})
        else:
            return jsonify({'success': False, 'error': 'Rules file not found'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/get_temp_signinout_data')
def get_temp_signinout_data():
    path = os.path.join(app.config['UPLOAD_FOLDER'], 'temp_signinout.xlsx')
    if not os.path.exists(path):
        return jsonify({'columns': [], 'data': []})
    df = pd.read_excel(path)
    cols = list(df.columns)
    rows = df.fillna('').astype(str).values.tolist()
    return jsonify({'columns': cols, 'data': rows})

@app.route('/get_temp_apply_data')
def get_temp_apply_data():
    path = os.path.join(app.config['UPLOAD_FOLDER'], 'temp_apply.xlsx')
    if not os.path.exists(path):
        return jsonify({'columns': [], 'data': []})
    df = pd.read_excel(path)
    cols = list(df.columns)
    rows = df.fillna('').astype(str).values.tolist()
    return jsonify({'columns': cols, 'data': rows})

@app.route('/get_temp_otlieu_data')
def get_temp_otlieu_data():
    path = os.path.join(app.config['UPLOAD_FOLDER'], 'temp_otlieu.xlsx')
    if not os.path.exists(path):
        return jsonify({'columns': [], 'data': []})
    df = pd.read_excel(path)
    cols = list(df.columns)
    rows = df.fillna('').astype(str).values.tolist()
    return jsonify({'columns': cols, 'data': rows})

LIEU_FOLLOWUP_PATH = os.path.join(app.config['UPLOAD_FOLDER'], 'lieu_followup.xlsx')
if os.path.exists(LIEU_FOLLOWUP_PATH):
    lieu_followup_df = pd.read_excel(LIEU_FOLLOWUP_PATH)
else:
    lieu_followup_df = pd.DataFrame(columns=['Name', 'Lieu remain previous month'])

@app.route('/get_lieu_followup')
def get_lieu_followup():
    if not os.path.exists(LIEU_FOLLOWUP_PATH):
        if employee_list_df is not None and not employee_list_df.empty:
            df = employee_list_df[['Name']].copy()
            df['Lieu remain previous month'] = 0
            df.insert(0, 'STT', range(1, len(df) + 1))
            df.to_excel(LIEU_FOLLOWUP_PATH, index=False)
        else:
            df = pd.DataFrame(columns=['STT', 'Name', 'Lieu remain previous month'])
    else:
        df = pd.read_excel(LIEU_FOLLOWUP_PATH)
    cols = list(df.columns)
    rows = df.fillna('').astype(str).values.tolist()
    return jsonify({'columns': cols, 'data': rows})

@app.route('/import_lieu_followup', methods=['POST'])
def import_lieu_followup():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    try:
        df = pd.read_excel(file)
        # Chuẩn hóa cột
        if 'Name' not in df.columns:
            return jsonify({'error': 'Missing Name column'}), 400
        if 'Lieu remain previous month' not in df.columns:
            df['Lieu remain previous month'] = 0
        df = df[['Name', 'Lieu remain previous month']]
        df.insert(0, 'STT', range(1, len(df) + 1))
        df.to_excel(LIEU_FOLLOWUP_PATH, index=False)
        return jsonify({'success': True, 'rows': len(df)})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/add_lieu_followup_row', methods=['POST'])
def add_lieu_followup_row():
    try:
        data = request.json
        name = data.get('Name', '').strip()
        remain = data.get('Lieu remain previous month', 0)
        if not name:
            return jsonify({'error': 'Name required'}), 400
        if os.path.exists(LIEU_FOLLOWUP_PATH):
            df = pd.read_excel(LIEU_FOLLOWUP_PATH)
        else:
            df = pd.DataFrame(columns=['STT', 'Name', 'Lieu remain previous month'])
        df = df.append({'Name': name, 'Lieu remain previous month': remain}, ignore_index=True)
        df['STT'] = range(1, len(df) + 1)
        df.to_excel(LIEU_FOLLOWUP_PATH, index=False)
        return jsonify({'success': True, 'rows': len(df)})
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/update_lieu_followup_row', methods=['POST'])
def update_lieu_followup_row():
    try:
        data = request.json
        idx = int(data.get('index'))
        col = data.get('column')
        value = data.get('value')
        if os.path.exists(LIEU_FOLLOWUP_PATH):
            df = pd.read_excel(LIEU_FOLLOWUP_PATH)
        else:
            return jsonify({'error': 'No data'}), 400
        if 0 <= idx < len(df) and col in df.columns:
            df.at[idx, col] = value
            df.to_excel(LIEU_FOLLOWUP_PATH, index=False)
            return jsonify({'success': True})
        else:
            return jsonify({'error': 'Invalid index or column'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 400

@app.route('/delete_lieu_followup_row', methods=['POST'])
def delete_lieu_followup_row():
    try:
        data = request.json
        idx = int(data.get('index'))
        if os.path.exists(LIEU_FOLLOWUP_PATH):
            df = pd.read_excel(LIEU_FOLLOWUP_PATH)
        else:
            return jsonify({'error': 'No data'}), 400
        if 0 <= idx < len(df):
            df = df.drop(idx).reset_index(drop=True)
            df['STT'] = range(1, len(df) + 1)
            df.to_excel(LIEU_FOLLOWUP_PATH, index=False)
            return jsonify({'success': True})
        else:
            return jsonify({'error': 'Invalid index'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 400

# ==== TÍNH TOÁN OT LIEU BEFORE (HÀM CHUNG) ====
def calculate_otlieu_before():
    global rules
    
    path = os.path.join(app.config['UPLOAD_FOLDER'], 'temp_otlieu.xlsx')
    if not os.path.exists(path):
        return pd.DataFrame()
    df = pd.read_excel(path)

    # Load Lieu followup
    lieu_followup_path = os.path.join(app.config['UPLOAD_FOLDER'], 'lieu_followup.xlsx')
    if os.path.exists(lieu_followup_path):
        lieu_followup_df = pd.read_excel(lieu_followup_path)
    else:
        lieu_followup_df = pd.DataFrame(columns=['Name', 'Lieu remain previous month'])
    
    # Load rules if not loaded
    if rules is None or rules.empty:
        rules_path = os.path.join(app.config['UPLOAD_FOLDER'], 'rules.xlsx')
        if os.path.exists(rules_path):
            rules = pd.read_excel(rules_path)
            print(f"Loaded rules from file: {len(rules)} rows")
        else:
            rules = pd.DataFrame()
            print("Rules file not found, using empty DataFrame")
    else:
        print(f"Using existing rules: {len(rules)} rows")

    # Build remain map: {name: remain}
    lieu_remain_map = {}
    if 'Name' in lieu_followup_df.columns and 'Lieu remain previous month' in lieu_followup_df.columns:
        for _, r in lieu_followup_df.iterrows():
            name = str(r['Name']).strip()
            try:
                remain = float(r['Lieu remain previous month'])
            except:
                remain = 0.0
            lieu_remain_map[name] = remain

    ot_payment_types = [
        'Weekday Rate 150%',
        'Weekday-night Rate 200%',
        'Weekend Rate 200%',
        'Weekend-night Rate 270%',
        'Holiday Rate 300%',
        'Holiday-night Rate 390%',
    ]
    change_in_lieu_types = ot_payment_types.copy()
    for col in ot_payment_types:
        df['OT Payment: ' + col] = 0.0
    for col in change_in_lieu_types:
        df['Change in lieu: ' + col] = 0.0

    # Hàm xác định loại ngày
    def get_day_type(dt, holidays, special_weekends, special_workdays):
        if dt in holidays:
            return 'Holiday'
        if dt in special_workdays:
            return 'Weekday'
        if dt in special_weekends:
            return 'Weekend'
        if dt.weekday() >= 5:
            return 'Weekend'
        return 'Weekday'

    holidays = set()
    special_workdays = set()
    special_weekends = set()
    try:
        if rules is not None:
            if 'Holiday Date in This Year' in rules.columns:
                holidays = set(pd.to_datetime(rules['Holiday Date in This Year'], errors='coerce').dt.date.dropna())
            if 'Special Work Day' in rules.columns:
                special_workdays = set(pd.to_datetime(rules['Special Work Day'], errors='coerce').dt.date.dropna())
            if 'Special Weekend' in rules.columns:
                special_weekends = set(pd.to_datetime(rules['Special Weekend'], errors='coerce').dt.date.dropna())
    except: pass

    # --- BẮT ĐẦU LOGIC TÍNH OT RATES ---
    for idx, row in df.iterrows():
        emp_id = row['Emp ID'] if 'Emp ID' in row else None
        intern = is_intern(emp_id, employee_list_df)
        
        # Ưu tiên lấy từ cột 'Date', sau đó 'OT date', sau đó 'Lieu Date'
        ot_date = None
        if 'Date' in df.columns and pd.notna(row.get('Date', None)):
            try:
                ot_date = pd.to_datetime(row['Date']).date()
            except:
                pass
        elif 'OT date' in df.columns and pd.notna(row.get('OT date', None)):
            try:
                ot_date = pd.to_datetime(row['OT date']).date()
            except:
                pass
        elif 'Lieu Date' in df.columns and pd.notna(row.get('Lieu Date', None)):
            try:
                ot_date = pd.to_datetime(row['Lieu Date']).date()
            except:
                pass
        if ot_date is None:
            continue

        # Lấy OT From và OT To
        ot_from, ot_to = None, None
        for col in df.columns:
            if 'ot' in col.lower() and 'from' in col.lower():
                ot_from = row[col]
            if 'ot' in col.lower() and 'to' in col.lower():
                ot_to = row[col]
        if not ot_from or not ot_to or pd.isna(ot_from) or pd.isna(ot_to):
            continue
        try:
            t1 = pd.to_datetime(str(ot_from), format='%H:%M')
            t2 = pd.to_datetime(str(ot_to), format='%H:%M')
        except:
            continue
        if t2 < t1:
            t2 += pd.Timedelta(days=1)

        # Tính OT theo từng block thời gian
        cur = t1
        while cur < t2:
            hour = cur.hour + cur.minute / 60
            if 6 <= hour < 22:
                block_end = min(cur.replace(hour=22, minute=0), t2)
                block_type = 'day'
            else:
                if hour < 6:
                    next6 = cur.replace(hour=6, minute=0)
                    if next6 <= cur: next6 += pd.Timedelta(days=1)
                    block_end = min(next6, t2)
                else:
                    next22 = cur.replace(hour=22, minute=0)
                    if next22 <= cur: next22 += pd.Timedelta(days=1)
                    block_end = min(next22, t2)
                block_type = 'night'
            block_date = (ot_date if cur.day == t1.day else ot_date + pd.Timedelta(days=1))
            day_type = get_day_type(block_date, holidays, special_weekends, special_workdays)
            hours = (block_end - cur).total_seconds() / 3600

            if block_type == 'day':
                lunch_start = cur.replace(hour=12, minute=0)
                lunch_end = cur.replace(hour=13, minute=30)
                overlap = max(timedelta(0), min(block_end, lunch_end) - max(cur, lunch_start)).total_seconds() / 3600
                if overlap > 0:
                    hours -= overlap

            target_prefix = 'Change in lieu: ' if intern else 'OT Payment: '
            rate_map = {
                ('Weekday', 'day'): 'Weekday Rate 150%',
                ('Weekday', 'night'): 'Weekday-night Rate 200%',
                ('Weekend', 'day'): 'Weekend Rate 200%',
                ('Weekend', 'night'): 'Weekend-night Rate 270%',
                ('Holiday', 'day'): 'Holiday Rate 300%',
                ('Holiday', 'night'): 'Holiday-night Rate 390%',
            }
            rate_label = rate_map.get((day_type, block_type))
            if rate_label:
                col = target_prefix + rate_label
                df.at[idx, col] += hours
            cur = block_end


    # --- LOGIC TRỪ LIEU ---
    for idx, row in df.iterrows():
        name = row['Name'] if 'Name' in row else None
        if not name:
            continue
            
        # Lấy thông tin Lieu Sum (giờ nghỉ Lieu)
        lieu_sum_col = next((c for c in df.columns if re.search(r'lieu.*sum', c, re.I)), None)
        lieu_sum = 0.0
        if lieu_sum_col:
            try:
                lieu_sum = float(row[lieu_sum_col]) if pd.notna(row[lieu_sum_col]) and str(row[lieu_sum_col]).strip() != '' else 0.0
            except:
                lieu_sum = 0.0
                
        # Nếu có Lieu Sum > 0, trừ vào OT Payment theo thứ tự ưu tiên
        if lieu_sum > 0:
            lieu_remain_old = lieu_remain_map.get(name, 0.0)
            lieu_to_deduct = min(lieu_sum, lieu_remain_old)
            lieu_used = 0.0
            
            if lieu_to_deduct > 0:
                # Thứ tự ưu tiên: Weekday → Night → Weekend → Holiday
                ot_priority = [
                    ('OT Payment: Weekday Rate 150%', 1.5, 'Change in lieu: Weekday Rate 150%'),
                    ('OT Payment: Weekday-night Rate 200%', 2.0, 'Change in lieu: Weekday-night Rate 200%'),
                    ('OT Payment: Weekend Rate 200%', 2.0, 'Change in lieu: Weekend Rate 200%'),
                    ('OT Payment: Weekend-night Rate 270%', 2.7, 'Change in lieu: Weekend-night Rate 270%'),
                    ('OT Payment: Holiday Rate 300%', 3.0, 'Change in lieu: Holiday Rate 300%'),
                    ('OT Payment: Holiday-night Rate 390%', 3.9, 'Change in lieu: Holiday-night Rate 390%'),
                ]
                
                remain = lieu_to_deduct
                for ot_col, ratio, lieu_col in ot_priority:
                    if remain <= 0:
                        break
                        
                    ot_val = df.at[idx, ot_col] if ot_col in df.columns else 0.0
                    try:
                        ot_val = float(ot_val) if pd.notna(ot_val) and str(ot_val).strip() != '' else 0.0
                    except:
                        ot_val = 0.0
                        
                    if ot_val <= 0:
                        continue
                        
                    # Số giờ OT có thể dùng để đổi Lieu ở hệ số này
                    max_lieu_from_this = ot_val * ratio
                    lieu_from_this = min(remain, max_lieu_from_this)
                    
                    # Số giờ OT bị trừ = lieu_from_this / ratio
                    ot_deduct = lieu_from_this / ratio
                    
                    # Cập nhật vào DataFrame
                    df.at[idx, ot_col] = round(ot_val - ot_deduct, 3)
                    df.at[idx, lieu_col] = round(lieu_from_this, 3)
                    
                    remain -= lieu_from_this
                    lieu_used += lieu_from_this
                
                # Cập nhật Lieu remain mới
                lieu_remain_new = round(lieu_remain_old - lieu_used, 3)
                lieu_remain_map[name] = lieu_remain_new
                
                # Ghi nhận vào DataFrame
                df.at[idx, 'Lieu used'] = round(lieu_used, 3)
                df.at[idx, 'Lieu Remain'] = round(lieu_remain_new, 3)
            else:
                df.at[idx, 'Lieu used'] = 0.0
                df.at[idx, 'Lieu Remain'] = round(lieu_remain_old, 3)
        else:
            # Không có Lieu Sum, ghi nhận Lieu Remain cũ
            lieu_remain_old = lieu_remain_map.get(name, 0.0)
            df.at[idx, 'Lieu used'] = 0.0
            df.at[idx, 'Lieu Remain'] = round(lieu_remain_old, 3)

    # Làm tròn 3 số
    for col in ['OT Payment: ' + c for c in ot_payment_types] + ['Change in lieu: ' + c for c in change_in_lieu_types]:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: round(float(x), 3) if str(x).replace('.', '', 1).replace('-', '', 1).isdigit() else x)

    # Ensure required columns exist and reorder
    import numpy as np
    if 'Lieu used' not in df.columns:
        df['Lieu used'] = np.nan
    if 'Lieu Remain' not in df.columns:
        df['Lieu Remain'] = np.nan
    if 'Remark' not in df.columns:
        df['Remark'] = np.nan
        
    # Reorder columns: put Lieu used, Lieu Remain, and Remark at the end
    cols = [c for c in df.columns if c not in ['Lieu used', 'Lieu Remain', 'Remark']] + ['Lieu used', 'Lieu Remain', 'Remark']
    df = df[cols]
    return df

@app.route('/get_otlieu_before')
def get_otlieu_before():
    df = calculate_otlieu_before()
    if df.empty:
        return jsonify({'columns': [], 'data': []})
    
    # Convert 0.0 values to empty strings for numeric columns
    numeric_cols = [col for col in df.columns if any(x in col for x in ['OT Payment:', 'Change in lieu:', 'Lieu used', 'Lieu Remain'])]
    
    for col in numeric_cols:
        if col in df.columns:
            df[col] = df[col].apply(lambda x: '' if pd.isna(x) or (isinstance(x, (int, float)) and x == 0.0) else x)
    
    cols = list(df.columns)
    rows = df.fillna('').astype(str).values.tolist()
    return jsonify({'columns': cols, 'data': rows})

@app.route('/get_otlieu_report')
def get_otlieu_report():
    emp_path = EMPLOYEE_LIST_PATH
    if os.path.exists(emp_path):
        emp_df = pd.read_csv(emp_path, dtype=str)
    else:
        emp_df = pd.DataFrame(columns=["Name", "ID Number"])
    emp_df = emp_df.fillna('')
    emp_df['No'] = range(1, len(emp_df) + 1)

    ot_df = calculate_otlieu_before()
    ot_cols = [
        'OT Payment: Weekday Rate 150%',
        'OT Payment: Weekday-night Rate 200%',
        'OT Payment: Weekend Rate 200%',
        'OT Payment: Weekend-night Rate 270%',
        'OT Payment: Holiday Rate 300%',
        'OT Payment: Holiday-night Rate 390%',
        'Change in lieu: Weekday Rate 150%',
        'Change in lieu: Weekday-night Rate 200%',
        'Change in lieu: Weekend Rate 200%',
        'Change in lieu: Weekend-night Rate 270%',
        'Change in lieu: Holiday Rate 300%',
        'Change in lieu: Holiday-night Rate 390%',
    ]
    ot_sum = None
    if not ot_df.empty and 'Name' in ot_df.columns:
        ot_df[ot_cols] = ot_df[ot_cols].fillna(0)
        ot_sum = ot_df.groupby('Name')[ot_cols].sum().reset_index()
        
        # Tính các cột mới theo yêu cầu
        used_cols = [c for c in ot_cols if c.startswith('Change in lieu')]
        paid_cols = [c for c in ot_cols if c.startswith('OT Payment')]
        
        # 1. Total used hours in month: Cộng lại các giờ đã dùng
        ot_sum['Total used hours in month'] = ot_sum[used_cols].sum(axis=1)
        
        # 2. Total OT paid: Cộng lại các cột OT Payment
        ot_sum['Total OT paid'] = ot_sum[paid_cols].sum(axis=1)
        
        # 3. Transferred to normal working hours: Logic mới
        def calc_transfer(row):
            try:
                total_ot_paid = float(row['Total OT paid'])
                total_used = float(row['Total used hours in month'])
                
                # Nếu có Lieu Sum (total_used) mà không có OT Payment (total_ot_paid = 0)
                # thì chuyển Lieu Sum vào "Transferred to normal working hours"
                if total_used > 0 and total_ot_paid == 0:
                    return round(total_used, 2)
                # Nếu có OT Payment > 25, thì chuyển phần dư vào normal working hours
                elif total_ot_paid > 25:
                    return round(total_ot_paid - 25, 2)
                else:
                    return ''
            except:
                return ''
        
        ot_sum['Transferred to normal working hours'] = ot_sum.apply(calc_transfer, axis=1)
        
        # 4. Remain unused time off in lieu: Phần Lieu chưa dùng hết
        # Load Lieu followup để lấy Lieu remain
        lieu_followup_path = os.path.join(app.config['UPLOAD_FOLDER'], 'lieu_followup.xlsx')
        lieu_remain_map = {}
        if os.path.exists(lieu_followup_path):
            lieu_followup_df = pd.read_excel(lieu_followup_path)
            if 'Name' in lieu_followup_df.columns and 'Lieu remain previous month' in lieu_followup_df.columns:
                for _, r in lieu_followup_df.iterrows():
                    name = str(r['Name']).strip()
                    try:
                        remain = float(r['Lieu remain previous month'])
                    except:
                        remain = 0.0
                    lieu_remain_map[name] = remain
        
        def calc_remain_unused(row):
            try:
                name = str(row['Name']).strip()
                lieu_remain_old = lieu_remain_map.get(name, 0.0)
                lieu_used = 0.0
                
                # Tính Lieu used từ các cột Change in lieu
                for col in used_cols:
                    try:
                        lieu_used += float(row[col]) if pd.notna(row[col]) else 0.0
                    except:
                        pass
                
                # Remain unused = Lieu remain cũ - Lieu used
                remain_unused = lieu_remain_old - lieu_used
                return round(remain_unused, 2) if remain_unused > 0 else 0.0
            except:
                return 0.0
        
        ot_sum['Remain unused time off in lieu'] = ot_sum.apply(calc_remain_unused, axis=1)
        ot_sum['Date'] = ''

    # Đảm bảo cột Name tồn tại ở cả hai DataFrame trước khi merge
    result = emp_df[['No', 'ID Number', 'Name']].copy()
    result = result.rename(columns={'ID Number': 'Employee ID'})
    # Không đổi tên 'Name' thành 'Employee Name' trước khi merge
    if ot_sum is not None:
        # Ép kiểu và strip để tránh lỗi merge do khác kiểu
        result['Name'] = result['Name'].astype(str).str.strip()
        ot_sum['Name'] = ot_sum['Name'].astype(str).str.strip()
        result = result.merge(ot_sum, on='Name', how='left')
    # Sau khi merge xong mới đổi tên cột cho thân thiện
    result = result.rename(columns={'Name': 'Employee Name'})
    col_rename = {
        'OT Payment: Weekday Rate 150%': 'OT weekday 150%',
        'OT Payment: Weekday-night Rate 200%': 'OT weekday night 200%',
        'OT Payment: Weekend Rate 200%': 'OT weekly holiday 200%',
        'OT Payment: Weekend-night Rate 270%': 'OT weekly holiday night 270%',
        'OT Payment: Holiday Rate 300%': 'OT public holiday 300%',
        'OT Payment: Holiday-night Rate 390%': 'OT public holiday night 390%',
        'Change in lieu: Weekday Rate 150%': 'OT weekday 150% (lieu)',
        'Change in lieu: Weekday-night Rate 200%': 'OT weekday night 200% (lieu)',
        'Change in lieu: Weekend Rate 200%': 'OT weekly holiday 200% (lieu)',
        'Change in lieu: Weekend-night Rate 270%': 'OT weekly holiday night 270% (lieu)',
        'Change in lieu: Holiday Rate 300%': 'OT public holiday 300% (lieu)',
        'Change in lieu: Holiday-night Rate 390%': 'OT public holiday night 390% (lieu)',
    }
    result = result.rename(columns=col_rename)
    col_order = [
        'No', 'Employee ID', 'Employee Name',
        'OT weekday 150%', 'OT weekday night 200%', 'OT weekly holiday 200%',
        'OT weekly holiday night 270%', 'OT public holiday 300%', 'OT public holiday night 390%',
        'OT weekday 150% (lieu)', 'OT weekday night 200% (lieu)', 'OT weekly holiday 200% (lieu)',
        'OT weekly holiday night 270% (lieu)', 'OT public holiday 300% (lieu)', 'OT public holiday night 390% (lieu)',
        'Transferred to normal working hours', 'Date', 'Total used hours in month',
        'Remain unused time off in lieu', 'Total OT paid'
    ]
    result = result[[c for c in col_order if c in result.columns]]
    
    # Convert 0.0 values to empty strings for numeric columns
    numeric_cols = [
        'OT weekday 150%', 'OT weekday night 200%', 'OT weekly holiday 200%',
        'OT weekly holiday night 270%', 'OT public holiday 300%', 'OT public holiday night 390%',
        'OT weekday 150% (lieu)', 'OT weekday night 200% (lieu)', 'OT weekly holiday 200% (lieu)',
        'OT weekly holiday night 270% (lieu)', 'OT public holiday 300% (lieu)', 'OT public holiday night 390% (lieu)',
        'Total used hours in month', 'Remain unused time off in lieu', 'Total OT paid'
    ]
    
    for col in numeric_cols:
        if col in result.columns:
            result[col] = result[col].apply(lambda x: '' if pd.isna(x) or (isinstance(x, (int, float)) and x == 0.0) else x)
    
    cols = list(result.columns)
    rows = result.fillna('').astype(str).values.tolist()
    return jsonify({'columns': cols, 'data': rows})

@app.route('/get_total_attendance_detail')
def get_total_attendance_detail():
    global _total_attendance_cache
    
    # Create cache key based on file modification times
    cache_key = f"total_attendance_{_get_cache_key()}"
    current_time = datetime.now().timestamp()
    
    # Check if cached result is still valid
    if cache_key in _total_attendance_cache:
        cached_data, timestamp = _total_attendance_cache[cache_key]
        if current_time - timestamp < _cache_timeout:
            return jsonify(cached_data)
    
    emp_path = EMPLOYEE_LIST_PATH
    if os.path.exists(emp_path):
        emp_df = pd.read_csv(emp_path, dtype=str)
    else:
        emp_df = pd.DataFrame(columns=["Name", "ID Number", "Dept"])

    emp_df = emp_df.fillna('')
    emp_df['No'] = range(1, len(emp_df) + 1)

    # Đổi tên các cột theo mẫu hiển thị
    result = emp_df.rename(columns={
        'ID Number': '14 Digits Employee ID',
        'Name': "Employee's name",
        'Dept': 'Group'
    })

    # Sắp xếp theo Group nếu có
    if 'Group' in result.columns:
        result = result.sort_values(by=['Group', 'No'], ascending=[True, True])
        result = result.reset_index(drop=True)
        result['No'] = range(1, len(result) + 1)

    # Khởi tạo các cột tính toán với giá trị 0
    attendance_cols = [
        'Normal working days',
        'Annual leave (100% salary)',
        'Sick leave (50% salary)',
        'Unpaid leave (0% salary)',
        'Welfare leave (100% salary)',
        'Total',
        'Late/Leave early (mins)',
        'Late/Leave early (times)',
        'Forget scanning',
        'Violation',
        'Remark',
        'Attendance for salary payment'
    ]
    for col in attendance_cols:
        result[col] = 0.0

    # Load và pre-process tất cả dữ liệu một lần
    signinout_data, apply_data, ot_lieu_data = [], [], []
    
    if os.path.exists(TEMP_SIGNINOUT_PATH):
        signinout_df = pd.read_excel(TEMP_SIGNINOUT_PATH)
        signinout_data = signinout_df.to_dict('records')
    
    if os.path.exists(TEMP_APPLY_PATH):
        apply_df = pd.read_excel(TEMP_APPLY_PATH)
        apply_data = apply_df.to_dict('records')
    
    if os.path.exists(TEMP_OTLIEU_PATH):
        otlieu_df = pd.read_excel(TEMP_OTLIEU_PATH)
        ot_lieu_data = otlieu_df.to_dict('records')

    # Lấy thông tin ngày đặc biệt từ rules
    holidays, special_weekends, special_workdays = get_special_days_from_rules(rules)

    # Xác định khoảng thời gian tính toán (19 tháng trước đến 20 tháng này)
    today = datetime.now()
    if today.day >= 19:
        start_date = today.replace(day=19) - timedelta(days=30)
        end_date = today.replace(day=20)
    else:
        start_date = today.replace(day=19) - timedelta(days=60)
        end_date = today.replace(day=20) - timedelta(days=30)
    
    date_range = pd.date_range(start=start_date, end=end_date, freq='D')

    # Helper functions
    def normalize_name(name_field):
        if not isinstance(name_field, str):
            return ""
        name_only = re.sub(r'\d{8,}$', '', name_field).strip()
        return name_only.lower()

    def get_day_type(dt, holidays, special_weekends, special_workdays):
        dt_date = dt.date()
        if dt_date in holidays:
            return 'Holiday'
        if dt_date in special_workdays:
            return 'Weekday'
        if dt_date in special_weekends:
            return 'Weekend'
        if dt.weekday() >= 5:
            return 'Weekend'
        return 'Weekday'

    def extract_name_from_emp_name(emp_name):
        import re
        if not isinstance(emp_name, str):
            return ""
        match = re.match(r'^(.+?)(\d{7,10})$', emp_name.strip())
        if match:
            return match.group(1).strip()
        return emp_name.strip()

    # Pre-process signin/out data - chỉ tạo DataFrame một lần
    df_sign_all = pd.DataFrame()
    if signinout_data:
        df_sign_all = pd.DataFrame(signinout_data)
        if 'emp_name' in df_sign_all.columns and 'attendance_time' in df_sign_all.columns:
            df_sign_all['attendance_time'] = pd.to_datetime(df_sign_all['attendance_time'], errors='coerce')
            df_sign_all['date'] = df_sign_all['attendance_time'].dt.date
            df_sign_all['target_name'] = df_sign_all['emp_name'].apply(extract_name_from_emp_name)

    # Pre-process apply data cho lookup nhanh hơn
    apply_lookup = {}
    for record in apply_data:
        if (record.get('Type') == 'Leave' and record.get('Results') == 'Approved'):
            target_name = extract_name_from_emp_name(record.get('Name', ''))
            if target_name not in apply_lookup:
                apply_lookup[target_name] = []
            apply_lookup[target_name].append(record)

    # Pre-process ot/lieu data cho lookup nhanh hơn
    ot_lieu_lookup = {}
    for record in ot_lieu_data:
        target_name = extract_name_from_emp_name(record.get('Name', ''))
        if target_name not in ot_lieu_lookup:
            ot_lieu_lookup[target_name] = []
        ot_lieu_lookup[target_name].append(record)

    # Optimized lookup functions
    def is_lieu_day(target_name, check_date):
        records = ot_lieu_lookup.get(target_name, [])
        for record in records:
            lieu_cols = ['Lieu From', 'Lieu To', 'Lieu From 2', 'Lieu To 2']
            for col in lieu_cols:
                if col in record and pd.notna(record[col]) and str(record[col]).strip():
                    try:
                        lieu_date = pd.to_datetime(record[col]).date()
                        if lieu_date == check_date:
                            return True
                    except:
                        continue
        return False

    def has_ot_on_date(target_name, check_date):
        records = ot_lieu_lookup.get(target_name, [])
        for record in records:
            ot_date_cols = ['OT Date', 'Date', 'OT date']
            for col in ot_date_cols:
                if col in record and pd.notna(record[col]) and str(record[col]).strip():
                    try:
                        ot_date = pd.to_datetime(record[col]).date()
                        if ot_date == check_date:
                            return True
                    except:
                        continue
        return False

    def has_apply_leave_on_date(target_name, check_date):
        records = apply_lookup.get(target_name, [])
        for record in records:
            try:
                start_date_str = record.get('Start Date', '')
                end_date_str = record.get('End Date', '')
                apply_date_str = record.get('Apply Date', '')
                
                if start_date_str:
                    start_date = pd.to_datetime(start_date_str).date()
                elif apply_date_str:
                    start_date = pd.to_datetime(apply_date_str).date()
                else:
                    continue
                    
                end_date = pd.to_datetime(end_date_str).date() if end_date_str else start_date
                
                if start_date <= check_date <= end_date:
                    return True
            except:
                continue
        return False

    def get_shift_info(target_name, check_date):
        if df_sign_all.empty:
            return 'NONE'
            
        # Filter records for this employee and date
        mask = (df_sign_all['target_name'] == target_name) & (df_sign_all['date'] == check_date)
        day_records = df_sign_all[mask]
        
        if day_records.empty:
            return 'NONE'
            
        # Process shift info
        morning_records = day_records[day_records['attendance_time'].dt.hour < 12]
        afternoon_records = day_records[day_records['attendance_time'].dt.hour >= 12]
        
        has_morning = not morning_records.empty
        has_afternoon = not afternoon_records.empty
        
        if has_morning and has_afternoon:
            return 'FULL'
        elif has_morning:
            return 'AM'
        elif has_afternoon:
            return 'PM'
        else:
            return 'NONE'

    # Tính toán cho từng nhân viên - tối ưu hóa bằng cách batch process
    for idx, emp in result.iterrows():
        emp_name = emp["Employee's name"]
        target_name = extract_name_from_emp_name(emp_name)
        
        normal_days = 0
        annual_leave = 0
        sick_leave = 0
        unpaid_leave = 0
        welfare_leave = 0
        late_early_mins = 0
        late_early_times = 0
        forget_scanning = 0
        violation = 0

        # Get all working days for this period
        workdays = []
        for dt in date_range:
            day_type = get_day_type(dt, holidays, special_weekends, special_workdays)
            dt_date = dt.date()
            
            is_normal_workday = (
                (day_type == 'Weekday') or
                (day_type == 'Weekend' and dt_date in special_workdays)
            )
            
            if is_normal_workday:
                workdays.append((dt, dt_date))

        # Batch check all conditions for workdays
        for dt, dt_date in workdays:
            has_ot = has_ot_on_date(target_name, dt_date)
            has_lieu = is_lieu_day(target_name, dt_date)
            has_apply_leave = has_apply_leave_on_date(target_name, dt_date)
            shift_info = get_shift_info(target_name, dt_date)
            
            # Calculate normal working days
            if not has_ot and not has_lieu and not has_apply_leave:
                if shift_info == 'FULL':
                    normal_days += 1.0
                elif shift_info in ['AM', 'PM']:
                    normal_days += 0.5
            
            # Calculate violations for working days
            if not df_sign_all.empty:
                mask = (df_sign_all['target_name'] == target_name) & (df_sign_all['date'] == dt_date)
                day_records = df_sign_all[mask]

                if day_records.empty:
                    forget_scanning += 1
                else:
                    morning_records = day_records[day_records['attendance_time'].dt.hour < 12]
                    afternoon_records = day_records[day_records['attendance_time'].dt.hour >= 12]
                    
                    in_time = morning_records['attendance_time'].min() if not morning_records.empty else None
                    out_time = afternoon_records['attendance_time'].max() if not afternoon_records.empty else None
                    
                    if pd.notna(in_time) and pd.notna(out_time):
                        # Check late arrival
                        if in_time.hour > 8 or (in_time.hour == 8 and in_time.minute > 30):
                            late_minutes = (in_time.hour - 8) * 60 + (in_time.minute - 30)
                            late_early_mins += late_minutes
                            late_early_times += 1
                        # Check early departure
                        if out_time.hour < 17 or (out_time.hour == 17 and out_time.minute < 30):
                            early_minutes = (17 - out_time.hour) * 60 + (30 - out_time.minute)
                            late_early_mins += early_minutes
                            late_early_times += 1

        # Process leave applications more efficiently
        records = apply_lookup.get(target_name, [])
        for record in records:
            try:
                start_date_str = record.get('Start Date', '')
                end_date_str = record.get('End Date', '')
                apply_date_str = record.get('Apply Date', '')
                
                if start_date_str:
                    start_date = pd.to_datetime(start_date_str).date()
                elif apply_date_str:
                    start_date = pd.to_datetime(apply_date_str).date()
                else:
                    continue
                    
                end_date = pd.to_datetime(end_date_str).date() if end_date_str else start_date
                
                # Count leave days in date range
                current_date = start_date
                while current_date <= end_date:
                    if any(d.date() == current_date for d in date_range):
                        day_type = get_day_type(pd.to_datetime(current_date), holidays, special_weekends, special_workdays)
                        if day_type == 'Weekday':
                            leave_type = str(record.get('Leave Type', '')).lower()
                            note = str(record.get('Note', '')).lower()
                            
                            # Determine leave duration
                            if any(keyword in note for keyword in ['morning', 'sáng', '上午', 'am']):
                                leave_days = 0.5
                            elif any(keyword in note for keyword in ['afternoon', 'chiều', '下午', 'pm']):
                                leave_days = 0.5
                            elif start_date == end_date:
                                leave_days = 1.0
                            else:
                                leave_days = 1.0
                            
                            # Categorize leave type
                            if 'annual' in leave_type:
                                annual_leave += leave_days
                            elif 'sick' in leave_type:
                                sick_leave += leave_days
                            elif 'unpaid' in leave_type:
                                unpaid_leave += leave_days
                            elif 'welfare' in leave_type:
                                welfare_leave += leave_days
                    
                    current_date += timedelta(days=1)
            except:
                continue

        # Update result
        result.at[idx, 'Normal working days'] = normal_days
        result.at[idx, 'Annual leave (100% salary)'] = annual_leave
        result.at[idx, 'Sick leave (50% salary)'] = sick_leave
        result.at[idx, 'Unpaid leave (0% salary)'] = unpaid_leave
        result.at[idx, 'Welfare leave (100% salary)'] = welfare_leave
        result.at[idx, 'Late/Leave early (mins)'] = late_early_mins
        result.at[idx, 'Late/Leave early (times)'] = late_early_times
        result.at[idx, 'Forget scanning'] = forget_scanning
        result.at[idx, 'Violation'] = violation
        
        # Calculate totals
        total_leave = annual_leave + sick_leave + unpaid_leave + welfare_leave
        result.at[idx, 'Total'] = total_leave
        result.at[idx, 'Attendance for salary payment'] = normal_days + total_leave

    # Reorder columns
    col_order = [
        'No', '14 Digits Employee ID',"Employee's name",'Group','Normal working days','Annual leave (100% salary)',
        'Sick leave (50% salary)','Unpaid leave (0% salary)','Welfare leave (100% salary)','Total',
        'Late/Leave early (mins)','Late/Leave early (times)','Forget scanning','Violation','Remark','Attendance for salary payment'
    ]
    result = result[[c for c in col_order if c in result.columns]]

    # Convert 0.0 and 0 values to empty strings for numeric columns
    numeric_cols = [
        'Normal working days', 'Annual leave (100% salary)', 'Sick leave (50% salary)',
        'Unpaid leave (0% salary)', 'Welfare leave (100% salary)', 'Total',
        'Late/Leave early (mins)', 'Late/Leave early (times)', 'Forget scanning', 
        'Violation', 'Attendance for salary payment'
    ]
    
    for col in numeric_cols:
        if col in result.columns:
            result[col] = result[col].apply(lambda x: '' if pd.isna(x) or (isinstance(x, (int, float)) and x == 0.0) else x)

    cols = list(result.columns)
    rows = result.fillna('').astype(str).values.tolist()

    result_data = {'columns': cols, 'data': rows}
    
    # Cache the result
    _total_attendance_cache[cache_key] = (result_data, current_time)

    return jsonify(result_data)

@app.route('/get_attendance_report')
def get_attendance_report():
    global _attendance_report_cache
    
    # Get month/year parameters from request
    selected_month = request.args.get('month', type=int)
    selected_year = request.args.get('year', type=int)
    
    # Create cache key including month/year parameters
    cache_key = f"attendance_report_{_get_cache_key()}_{selected_month}_{selected_year}"
    current_time = datetime.now().timestamp()
    
    # Check if cached result is still valid
    if cache_key in _attendance_report_cache:
        cached_data, timestamp = _attendance_report_cache[cache_key]
        if current_time - timestamp < _cache_timeout:
            return jsonify(cached_data)
    
    # Lấy dữ liệu từ các file tạm trước
    signinout_data, apply_data, ot_lieu_data = [], [], []
    
    if os.path.exists(TEMP_SIGNINOUT_PATH):
        signinout_df = pd.read_excel(TEMP_SIGNINOUT_PATH)
        signinout_data = signinout_df.to_dict('records')
    
    if os.path.exists(TEMP_APPLY_PATH):
        apply_df = pd.read_excel(TEMP_APPLY_PATH)
        apply_data = apply_df.to_dict('records')
    
    if os.path.exists(TEMP_OTLIEU_PATH):
        otlieu_df = pd.read_excel(TEMP_OTLIEU_PATH)
        ot_lieu_data = otlieu_df.to_dict('records')

    # Determine month and year from parameters or auto-detect from data
    if selected_month and selected_year:
        # Use user-selected month/year
        month, year = selected_month, selected_year
        print(f"Using user-selected month/year: {month}/{year}")
    else:
        # Auto-detect from data (existing logic)
        month, year = 7, 2024
        if signinout_data:
            dates = [pd.to_datetime(r['attendance_time']) for r in signinout_data if pd.notna(r.get('attendance_time'))]
            if dates:
                month_counts = {}
                for date in dates:
                    month_key = (date.month, date.year)
                    month_counts[month_key] = month_counts.get(month_key, 0) + 1
                most_common_month = max(month_counts.items(), key=lambda x: x[1])[0]
                month, year = most_common_month
        print(f"Auto-detected month/year: {month}/{year}")

    # Load employee list
    emp_df = pd.read_csv(EMPLOYEE_LIST_PATH, dtype=str) if os.path.exists(EMPLOYEE_LIST_PATH) else pd.DataFrame(columns=["Dept", "Name"])
    if 'Dept' in emp_df.columns:
        emp_df = emp_df.sort_values(by=['Dept', 'Name']).reset_index(drop=True)

    # Calculate date range: 19th of previous month to 20th of current month
    if month == 1:
        prev_month = 12
        prev_year = year - 1
    else:
        prev_month = month - 1
        prev_year = year
    
    start_date = pd.Timestamp(prev_year, prev_month, 19)
    end_date = pd.Timestamp(year, month, 20)
    days = pd.date_range(start=start_date, end=end_date, freq='D')
    day_cols = [d.strftime('%Y-%m-%d') for d in days]
    
    holidays, special_weekends, special_workdays = get_special_days_from_rules(rules)
    
    # Helper functions - define before use
    def normalize_name(name_field):
        import re
        if not isinstance(name_field, str):
            return ""
        name_only = re.sub(r'\d{8,}$', '', name_field).strip()
        return name_only.lower()

    def get_day_type(dt, holidays, special_weekends, special_workdays):
        dt_date = dt.date()
        if dt_date in holidays:
            return 'Holiday'
        if dt_date in special_workdays:
            return 'Weekday'
        if dt_date in special_weekends:
            return 'Weekend'
        if dt.weekday() >= 5:  # Saturday = 5, Sunday = 6
            return 'Weekend'
        return 'Weekday'
    
    # Optimize: Pre-process signin/out data once
    df_sign_all = pd.DataFrame()
    if signinout_data:
        df_sign_all = pd.DataFrame(signinout_data)
        if 'emp_name' in df_sign_all.columns and 'attendance_time' in df_sign_all.columns:
            df_sign_all['attendance_time'] = pd.to_datetime(df_sign_all['attendance_time'], errors='coerce')
            df_sign_all['Date'] = df_sign_all['attendance_time'].dt.date
            df_sign_all['NormalizedName'] = df_sign_all['emp_name'].apply(normalize_name)

    def extract_name_from_emp_name(emp_name):
        import re
        if not isinstance(emp_name, str):
            return ""
        match = re.match(r'^(.+?)(\d{7,10})$', emp_name.strip())
        if match:
            return match.group(1).strip()
        return emp_name.strip()

    # Pre-process apply data for faster lookup
    apply_lookup = {}
    for record in apply_data:
        if record.get('Results') == 'Approved':
            name_val = str(record.get('Name', '') or '')
            normalized_name_val = normalize_name(name_val)
            if normalized_name_val not in apply_lookup:
                apply_lookup[normalized_name_val] = []
            apply_lookup[normalized_name_val].append(record)

    # Pre-process OT/Lieu data for faster lookup
    ot_lieu_lookup = {}
    for record in ot_lieu_data:
        name_val = str(record.get('Name', '') or '')
        normalized_name_val = normalize_name(name_val)
        if normalized_name_val not in ot_lieu_lookup:
            ot_lieu_lookup[normalized_name_val] = []
        ot_lieu_lookup[normalized_name_val].append(record)

    # Optimized lookup functions using pre-fetched data
    def _is_lieu_day_optimized(lieu_records, check_date):
        """Optimized lieu day check with pre-fetched records"""
        for record in lieu_records:
            lieu_cols = ['Lieu From', 'Lieu To', 'Lieu From 2', 'Lieu To 2']
            for col in lieu_cols:
                if col in record and pd.notna(record[col]) and str(record[col]).strip():
                    try:
                        lieu_date = pd.to_datetime(record[col]).date()
                        if lieu_date == check_date:
                            return True
                    except:
                        continue
        return False

    def _get_apply_info_optimized(apply_records, check_date):
        """Optimized apply info lookup with pre-fetched records"""
        for record in apply_records:
            if record.get('Results') != 'Approved':
                continue
            try:
                start_date_str = record.get('Start Date', '')
                end_date_str = record.get('End Date', '')
                apply_date_str = record.get('Apply Date', '')
                
                if start_date_str:
                    start_date = pd.to_datetime(start_date_str).date()
                elif apply_date_str:
                    start_date = pd.to_datetime(apply_date_str).date()
                else:
                    continue
                    
                end_date = pd.to_datetime(end_date_str).date() if end_date_str else start_date
                
                if start_date <= check_date <= end_date:
                    return {
                        'type': record.get('Type', ''),
                        'leave_type': record.get('Leave Type', ''),
                        'is_approved': True
                    }
            except:
                continue
        return None

    def _calculate_status_optimized(record, lieu_records):
        """Optimized status calculation"""
        if record['DayType'] != 'Weekday' or record['IsLieu']:
            if record['IsLieu']:
                # Simplified lieu processing
                return 'Lieu', _calculate_lieu_penalty(record, lieu_records)
            else:
                return 'Off', 0
        elif record['ApplyInfo'] and record['ApplyInfo']['is_approved']:
            return _process_apply_status(record)
        elif pd.isna(record.get('SignIn')) or pd.isna(record.get('SignOut')):
            return 'Miss', 0
        else:
            return _process_normal_status(record)

    def _calculate_lieu_penalty(record, lieu_records):
        """Simplified lieu penalty calculation"""
        try:
            # Quick penalty calculation without complex validation
            lieu_hours = 0
            for lieu_record in lieu_records:
                lieu_cols = ['Lieu From', 'Lieu To', 'Lieu From 2', 'Lieu To 2']
                for i in range(0, len(lieu_cols), 2):
                    lieu_from_col = lieu_cols[i]
                    lieu_to_col = lieu_cols[i+1] if i+1 < len(lieu_cols) else None
                    
                    if (lieu_from_col in lieu_record and lieu_to_col in lieu_record and
                        pd.notna(lieu_record[lieu_from_col]) and pd.notna(lieu_record[lieu_to_col])):
                        try:
                            lieu_from = pd.to_datetime(f"{record['Date']} {lieu_record[lieu_from_col]}")
                            lieu_to = pd.to_datetime(f"{record['Date']} {lieu_record[lieu_to_col]}")
                            lieu_duration = (lieu_to - lieu_from).total_seconds() / 3600
                            lieu_hours += lieu_duration
                        except:
                            continue
            
            # Simple work hours calculation
            work_hours = 0
            if pd.notna(record.get('SignIn')) and pd.notna(record.get('SignOut')):
                work_seconds = (record['SignOut'] - record['SignIn']).total_seconds()
                # Subtract lunch break if spans lunch time
                if (record['SignIn'].hour < 13 and record['SignOut'].hour > 12):
                    work_seconds -= 1.5 * 3600
                work_hours = work_seconds / 3600
            
            total_hours = work_hours + lieu_hours
            if total_hours < 8.0:
                return int((8.0 - total_hours) * 60)
            return 0
        except:
            return 45  # Standard penalty for errors

    def _process_apply_status(record):
        """Simplified apply status processing"""
        apply_type = record['ApplyInfo']['type']
        if apply_type == 'Supp':
            return 'Supp', 0  # Simplified - no complex validation
        elif apply_type in ['Trip', 'Leave']:
            return apply_type, 0
        else:
            return apply_type if apply_type else 'Normal', 0

    def _process_normal_status(record):
        """Simplified normal status processing"""
        check_in_time, check_out_time = record['SignIn'], record['SignOut']
        
        if pd.notna(check_in_time) and pd.notna(check_out_time):
            # Simplified work hours calculation
            work_seconds = (check_out_time - check_in_time).total_seconds()
            
            # Subtract lunch if spans lunch period
            if check_in_time.hour < 13 and check_out_time.hour > 12:
                work_seconds -= 1.5 * 3600
            
            work_hours = work_seconds / 3600
            
            # Simple late/early check
            if (check_in_time.hour > 8 or (check_in_time.hour == 8 and check_in_time.minute > 30) or
                check_out_time.hour < 17 or (check_out_time.hour == 17 and check_out_time.minute < 30) or
                work_hours < 8.0):
                
                # Calculate penalty
                late_mins = max(0, (check_in_time.hour - 8) * 60 + check_in_time.minute - 30)
                early_mins = max(0, (17 - check_out_time.hour) * 60 + (30 - check_out_time.minute))
                shortage_mins = max(0, int((8.0 - work_hours) * 60))
                
                return 'Late/Soon', max(late_mins, early_mins, shortage_mins)
            else:
                return 'Normal', 0
        return 'Miss', 0

    # Optimized lookup functions using pre-fetched data
    def is_lieu_day(normalized_name, check_date):
        records = ot_lieu_lookup.get(normalized_name, [])
        return _is_lieu_day_optimized(records, check_date)

    def get_apply_info_for_date(normalized_name, check_date):
        records = apply_lookup.get(normalized_name, [])
        return _get_apply_info_optimized(records, check_date)

    # Prepare employee data with normalized names
    if os.path.exists(EMPLOYEE_LIST_PATH):
        emp_df = pd.read_csv(EMPLOYEE_LIST_PATH, dtype=str)
        emp_df.dropna(subset=['Name'], inplace=True)
        emp_df['NormalizedName'] = emp_df['Name'].apply(normalize_name)
        emp_df['DisplayName'] = emp_df['Name'].apply(lambda x: re.sub(r'\d{8,}$', '', x).strip())
    else:
        return jsonify({'error': 'Không tìm thấy file danh sách nhân viên.'})

    # Pre-aggregate daily attendance data with time constraints - OPTIMIZED
    def process_daily_attendance_with_constraints(df):
        """Process attendance with time constraints per specification - Vectorized"""
        if df.empty:
            return pd.DataFrame(columns=['NormalizedName', 'Date', 'MorningTime', 'AfternoonTime', 'SignIn', 'SignOut'])
        
        # Filter valid times once for all data
        df_valid = df[(df['attendance_time'].dt.hour >= 8) & (df['attendance_time'].dt.hour <= 18)].copy()
        
        if df_valid.empty:
            # Create empty result for all name-date combinations
            unique_combos = df[['NormalizedName', 'Date']].drop_duplicates()
            result = unique_combos.copy()
            result['MorningTime'] = None
            result['AfternoonTime'] = None 
            result['SignIn'] = None
            result['SignOut'] = None
            return result
        
        # Vectorized operations using groupby.agg
        agg_funcs = {
            'attendance_time': ['min', 'max']
        }
        
        # Overall SignIn/SignOut
        result = df_valid.groupby(['NormalizedName', 'Date']).agg(agg_funcs).reset_index()
        result.columns = ['NormalizedName', 'Date', 'SignIn', 'SignOut']
        
        # Morning times (< 4 PM) - vectorized
        morning_df = df_valid[df_valid['attendance_time'].dt.hour < 16]
        if not morning_df.empty:
            morning_times = morning_df.groupby(['NormalizedName', 'Date'])['attendance_time'].min().reset_index()
            morning_times.columns = ['NormalizedName', 'Date', 'MorningTime']
            result = result.merge(morning_times, on=['NormalizedName', 'Date'], how='left')
        else:
            result['MorningTime'] = None
        
        # Afternoon times (>= 10:30) - vectorized
        cutoff_time = pd.Timestamp('10:30').time()
        afternoon_df = df_valid[df_valid['attendance_time'].dt.time >= cutoff_time]
        if not afternoon_df.empty:
            afternoon_times = afternoon_df.groupby(['NormalizedName', 'Date'])['attendance_time'].max().reset_index()
            afternoon_times.columns = ['NormalizedName', 'Date', 'AfternoonTime']
            result = result.merge(afternoon_times, on=['NormalizedName', 'Date'], how='left')
        else:
            result['AfternoonTime'] = None
        
        return result
    
    if not df_sign_all.empty:
        daily_times = process_daily_attendance_with_constraints(df_sign_all)
    else:
        daily_times = pd.DataFrame(columns=['NormalizedName', 'Date', 'MorningTime', 'AfternoonTime', 'SignIn', 'SignOut'])

    # Create master dataframe for all employee-day combinations - OPTIMIZED
    # Use list comprehension and vectorized operations
    emp_names = emp_df['NormalizedName'].tolist()
    emp_display_names = emp_df['DisplayName'].tolist()
    emp_depts = emp_df.get('Dept', [''] * len(emp_df)).tolist()
    
    # Create cartesian product efficiently
    master_data = []
    for i, (norm_name, disp_name, dept) in enumerate(zip(emp_names, emp_display_names, emp_depts)):
        for day in days:
            master_data.append({
                'DisplayName': disp_name,
                'NormalizedName': norm_name,
                'Dept': dept,
                'Date': day.date(),
                'DayPandas': day
            })
    
    master_df = pd.DataFrame(master_data)
    
    # Merge with attendance data
    master_df = pd.merge(master_df, daily_times, on=['NormalizedName', 'Date'], how='left')

    # Pre-compute common values to avoid repeated calculations
    day_type_cache = {}
    for day in days:
        day_type_cache[day.date()] = get_day_type(day, holidays, special_weekends, special_workdays)

    processed_records = []
    abnormal_report_data = []

    # Vectorized processing - group by normalized name to reduce lookups
    for norm_name, emp_group in master_df.groupby('NormalizedName'):
        # Pre-fetch data for this employee once
        lieu_records = ot_lieu_lookup.get(norm_name, [])
        apply_records = apply_lookup.get(norm_name, [])
        
        # Process all records for this employee
        for _, row in emp_group.iterrows():
            emp_name, day_date, day_pd_ts = row['DisplayName'], row['Date'], row['DayPandas']

            record = {
                'DisplayName': emp_name,
                'NormalizedName': norm_name,
                'Date': day_date,
                'Dept': row['Dept']
            }

            # Use cached day type
            record['DayType'] = day_type_cache[day_date]
            
            # Optimized lieu check with pre-fetched data
            record['IsLieu'] = _is_lieu_day_optimized(lieu_records, day_date)
            
            # Optimized apply info with pre-fetched data
            record['ApplyInfo'] = _get_apply_info_optimized(apply_records, day_date)
            
            record['SignIn'] = row.get('SignIn')
            record['SignOut'] = row.get('SignOut')
            
            # Simplified status calculation
            status, late_minutes = _calculate_status_optimized(record, lieu_records)

            record['Status'] = status
            record['LateMinutes'] = late_minutes
            processed_records.append(record)

    processed_df = pd.DataFrame(processed_records)

    # Build result table - OPTIMIZED using vectorized operations
    result_rows = []
    summary_keys = ['Normal', 'Leave', 'Trip', 'Miss', 'Late/Soon', 'Lieu', 'Supp', 'TotalLateMinutes']

    # Pre-compute day strings to avoid repeated formatting
    day_strs = [day.strftime('%Y-%m-%d') for day in days]
    day_to_str = {day.date(): day_str for day, day_str in zip(days, day_strs)}

    for _, emp in emp_df.iterrows():
        emp_name_display = emp['DisplayName']
        emp_name_normalized = emp['NormalizedName']
        
        # Filter employee data once
        emp_data = processed_df[processed_df['NormalizedName'] == emp_name_normalized]
        
        # Pre-build employee rows
        morning_row = {'Department': emp.get('Dept', ''), 'Name': emp_name_display, 'Shift': 'Morning shift'}
        afternoon_row = {'Department': emp.get('Dept', ''), 'Name': emp_name_display, 'Shift': 'Afternoon shift'}
        
        summary = {key: 0 for key in summary_keys}

        # Process all days for this employee at once
        if not emp_data.empty:
            # Create a lookup for faster day record access
            day_records = emp_data.set_index('Date')
            
            for day in days:
                day_date = day.date()
                day_str = day_to_str[day_date]
                
                if day_date in day_records.index:
                    record = day_records.loc[day_date]
                    status = record['Status']
                    
                    if status in ['Trip', 'Leave', 'Supp', 'Miss']:
                        morning_row[day_str], afternoon_row[day_str] = status, status
                        summary[status] += 0.5
                    elif status == 'Lieu':
                        morning_row[day_str], afternoon_row[day_str] = 'Lieu', 'Lieu'
                        summary['Lieu'] += 0.5
                    elif status == 'Off':
                        morning_row[day_str], afternoon_row[day_str] = '', ''
                    elif status in ['Normal', 'Late/Soon']:
                        # Format times only when needed
                        morning_time = record['SignIn'].strftime('%H:%M') if pd.notna(record['SignIn']) else '0'
                        afternoon_time = record['SignOut'].strftime('%H:%M') if pd.notna(record['SignOut']) else '0'
                        
                        prefix = "LATE:" if status == 'Late/Soon' else "NORMAL:"
                        morning_row[day_str] = f"{prefix}{morning_time}"
                        afternoon_row[day_str] = f"{prefix}{afternoon_time}"
                        summary[status] += 0.5
                    
                    summary['TotalLateMinutes'] += record['LateMinutes']
                else:
                    # No record for this day
                    morning_row[day_str], afternoon_row[day_str] = '', ''
        else:
            # No data for this employee
            for day_str in day_strs:
                morning_row[day_str], afternoon_row[day_str] = '', ''

        # Add summary to rows
        morning_row.update(summary)
        afternoon_row.update(summary)
        
        # Add penalty tracking to afternoon row
        afternoon_row['TotalPenalty'] = summary['TotalLateMinutes']
        afternoon_row['HasRedHighlight'] = summary['TotalLateMinutes'] >= 30
        
        result_rows.append(morning_row)
        result_rows.append(afternoon_row)

    # Optimize final result building - use day_strs instead of day_cols
    result = pd.DataFrame(result_rows)
    columns = ['Department', 'Name', 'Shift'] + day_strs + list(summary_keys)
    
    # Ensure all columns exist and reorder efficiently
    missing_cols = set(columns) - set(result.columns)
    for col in missing_cols:
        result[col] = ''
    result = result[columns]
    
    # Convert to response format efficiently
    cols = result.columns.tolist()
    rows = result.values.tolist()  # Direct values conversion is faster than fillna().astype()

    # Build abnormal data more efficiently - OPTIMIZED
    abnormal_late_early_data = []
    abnormal_missing_data = []

    # Pre-filter workdays to avoid repeated calculations
    workdays = []
    for day in days:
        day_date = day.date()
        day_type = day_type_cache[day_date]
        if day_type == 'Weekday' or (day_type == 'Weekend' and day_date in special_workdays):
            workdays.append((day, day_date))

    # Process abnormal data using vectorized operations where possible
    for _, emp_row in emp_df.iterrows():
        emp_name = emp_row['DisplayName']
        emp_name_normalized = emp_row['NormalizedName']
        
        # Get all data for this employee at once
        emp_data = processed_df[processed_df['NormalizedName'] == emp_name_normalized]
        if emp_data.empty:
            continue
            
        # Create lookup for faster access
        emp_day_lookup = emp_data.set_index('Date')

        for day, day_date in workdays:
            if day_date in emp_day_lookup.index:
                record = emp_day_lookup.loc[day_date]
                check_in = record.get('SignIn')
                check_out = record.get('SignOut')

                if pd.isna(check_in) or pd.isna(check_out):
                    abnormal_missing_data.append({
                        'Name': emp_name,
                        'Date': day_date.strftime('%Y-%m-%d'),
                        'Check in': check_in.strftime('%H:%M') if pd.notna(check_in) else '',
                        'Check out': check_out.strftime('%H:%M') if pd.notna(check_out) else '',
                        'Need apply on TMS': 'Need apply on TMS'
                    })
                else:
                    # Quick calculation for late/early penalties
                    late_mins = max(0, (check_in.hour - 8) * 60 + check_in.minute - 30)
                    early_mins = max(0, (17 - check_out.hour) * 60 + (30 - check_out.minute))
                    total_penalty_mins = late_mins + early_mins

                    if total_penalty_mins > 0:
                        status = []
                        if late_mins > 0:
                            status.append('Late')
                        if early_mins > 0:
                            status.append('Early')

                        abnormal_late_early_data.append({
                            'Name': emp_name,
                            'Attendance Date': day_date.strftime('%Y-%m-%d'),
                            'morning card-swipe': check_in.strftime('%H:%M'),
                            'afternoon card-swipe': check_out.strftime('%H:%M'),
                            'status': ' & '.join(status),
                            'Unpaid salary due to late come/early leave (mins)': total_penalty_mins
                        })
            else:
                # No attendance record
                abnormal_missing_data.append({
                    'Name': emp_name,
                    'Date': day_date.strftime('%Y-%m-%d'),
                    'Check in': '',
                    'Check out': '',
                    'Need apply on TMS': 'Need apply on TMS'
                })

    # Convert abnormal data to response format efficiently
    late_early_cols = ['Name', 'Attendance Date', 'morning card-swipe', 'afternoon card-swipe', 'status', 'Unpaid salary due to late come/early leave (mins)']
    late_early_rows = [[d.get(col, '') for col in late_early_cols] for d in abnormal_late_early_data]

    missing_cols = ['Name', 'Date', 'Check in', 'Check out', 'Need apply on TMS']
    missing_rows = [[d.get(col, '') for col in missing_cols] for d in abnormal_missing_data]

    result_data = {
        'columns': cols, 'rows': rows,
        'abnormal_late_early_columns': late_early_cols,
        'abnormal_late_early_rows': late_early_rows,
        'abnormal_missing_columns': missing_cols,
        'abnormal_missing_rows': missing_rows,
        'month': month,  # Add month/year to response
        'year': year,
        'period_start': start_date.strftime('%Y-%m-%d'),
        'period_end': end_date.strftime('%Y-%m-%d')
    }
    
    # Cache the result
    _attendance_report_cache[cache_key] = (result_data, current_time)
    
    return jsonify(result_data)

@app.route('/get_available_months')
def get_available_months():
    """Get list of available months from imported data - based on attendance report period logic"""
    try:
        # Load sign-in/out data to check available months
        signinout_data = []
        if os.path.exists(TEMP_SIGNINOUT_PATH):
            signinout_df = pd.read_excel(TEMP_SIGNINOUT_PATH)
            signinout_data = signinout_df.to_dict('records')
        
        available_months = []
        
        if signinout_data:
            dates = [pd.to_datetime(r['attendance_time']) for r in signinout_data if pd.notna(r.get('attendance_time'))]
            if dates:
                # Map each date to its report month using attendance report logic
                # Report month logic: from 19th of previous month to 20th of current month
                report_month_counts = {}
                
                for date in dates:
                    # Determine which report month this date belongs to
                    if date.day >= 19:
                        # If day >= 19, belongs to next month's report
                        if date.month == 12:
                            report_month = 1
                            report_year = date.year + 1
                        else:
                            report_month = date.month + 1
                            report_year = date.year
                    else:
                        # If day < 19, belongs to current month's report
                        report_month = date.month
                        report_year = date.year
                    
                    month_key = (report_month, report_year)
                    report_month_counts[month_key] = report_month_counts.get(month_key, 0) + 1
                
                # Sort by year, then by month
                sorted_months = sorted(report_month_counts.keys(), key=lambda x: (x[1], x[0]))
                
                for month, year in sorted_months:
                    available_months.append({
                        'month': month,
                        'year': year,
                        'display': f"{month:02d}/{year}",
                        'record_count': report_month_counts[(month, year)]
                    })
        
        return jsonify({
            'success': True,
            'available_months': available_months,
            'current_month': datetime.now().month,
            'current_year': datetime.now().year
        })
        
    except Exception as e:
        return jsonify({
            'success': False,
            'error': str(e),
            'available_months': [],
            'current_month': datetime.now().month,
            'current_year': datetime.now().year
        })

def flatten_cell(cell):
    if isinstance(cell, dict) and 'value' in cell:
        return cell['value']
    return cell

def get_special_days_from_rules(rules):
    holidays = set()
    special_weekends = set()
    special_workdays = set()
    if rules is not None:
        if 'Holiday Date in This Year' in rules.columns:
            holidays = set(pd.to_datetime(rules['Holiday Date in This Year'], errors='coerce').dt.date.dropna())
        if 'Special Weekend' in rules.columns:
            special_weekends = set(pd.to_datetime(rules['Special Weekend'], errors='coerce').dt.date.dropna())
        if 'Special Work Day' in rules.columns:
            special_workdays = set(pd.to_datetime(rules['Special Work Day'], errors='coerce').dt.date.dropna())
    return holidays, special_weekends, special_workdays

def is_intern(emp_id, emp_list_df):
    if emp_id in emp_list_df['ID Number'].values:
        emp_row = emp_list_df[emp_list_df['ID Number'] == emp_id].iloc[0]
        return emp_row.get('Internship', '') == 'Intern'
    return False

# ========================
# MULTI-FILE IMPORT ROUTES
# ========================
@app.route('/import_multiple_files', methods=['POST'])
def import_multiple_files():
    """Import multiple files for a specific data type with append functionality"""
    global sign_in_out_data, apply_data, ot_lieu_data, employee_list_df
    
    data_type = request.form.get('data_type')
    if not data_type:
        return jsonify({'error': 'Data type is required'}), 400
    
    if 'files[]' not in request.files:
        return jsonify({'error': 'No files uploaded'}), 400
    
    files = request.files.getlist('files[]')
    if not files or all(f.filename == '' for f in files):
        return jsonify({'error': 'No files selected'}), 400
    
    total_rows = 0
    imported_files = []
    
    try:
        for file in files:
            if file.filename == '':
                continue
                
            if not file.filename.lower().endswith(('.xlsx', '.xls', '.csv')):
                continue
                
            filename = secure_filename(file.filename)
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            file.save(file_path)
            
            # Load data based on file type
            if file.filename.lower().endswith('.csv'):
                df = try_read_csv(open(file_path, 'rb').read())
            elif file.filename.lower().endswith('.xls'):
                try:
                    # Thử với xlrd engine trước
                    df = pd.read_excel(file_path, engine='xlrd')
                except Exception as e:
                    print(f"xlrd engine failed for .xls multiple import: {e}")
                    try:
                        # Fallback: thử với openpyxl engine
                        df = pd.read_excel(file_path, engine='openpyxl')
                    except Exception as e2:
                        print(f"openpyxl engine also failed for .xls multiple import: {e2}")
                        return jsonify({'error': f'Cannot read .xls file {filename}. Please convert to .xlsx format. Error: {str(e)}'}), 400
            else:
                df = pd.read_excel(file_path)
            
            df = df.dropna(how='all').fillna('')
            
            # Process data based on type
            if data_type == 'signinout':
                keep = [col for col in df.columns if col.lower() in ['emp_name', 'attendance_time']]
                df = df[keep]
                if sign_in_out_data is None or sign_in_out_data.empty:
                    sign_in_out_data = df
                else:
                    sign_in_out_data = pd.concat([sign_in_out_data, df], ignore_index=True)
                total_rows = len(sign_in_out_data)
                
            elif data_type == 'apply':
                if 'support' in filename.lower() or 'suport' in filename.lower():
                    df = transform_support_apply_df(df, employee_list_df)
                df = translate_apply_headers(df)
                df = filter_apply_employees(df, employee_list_df)
                df = add_apply_columns(df)
                if apply_data is None or apply_data.empty:
                    apply_data = df
                else:
                    apply_data = pd.concat([apply_data, df], ignore_index=True)
                total_rows = len(apply_data)
                
            elif data_type == 'otlieu':
                if 'OT From Note: 12AM is midnight' in df.columns:
                    df = df.rename(columns={'OT From Note: 12AM is midnight': 'OT From'})
                df = process_ot_lieu_df(df, employee_list_df)
                if ot_lieu_data is None or ot_lieu_data.empty:
                    ot_lieu_data = df
                else:
                    ot_lieu_data = pd.concat([ot_lieu_data, df], ignore_index=True)
                total_rows = len(ot_lieu_data)
            
            imported_files.append(filename)
        
        return jsonify({
            'success': True,
            'message': f'Successfully imported {len(imported_files)} files. Total rows: {total_rows}',
            'imported_files': imported_files,
            'total_rows': total_rows
        })
        
    except Exception as e:
        return jsonify({'error': f'Error importing files: {str(e)}'}), 400

@app.route('/remove_file_from_data', methods=['POST'])
def remove_file_from_data():
    """Remove specific file data from the current dataset"""
    global sign_in_out_data, apply_data, ot_lieu_data
    
    data_type = request.form.get('data_type')
    file_index = request.form.get('file_index')
    
    if not data_type or file_index is None:
        return jsonify({'error': 'Data type and file index are required'}), 400
    
    try:
        file_index = int(file_index)
        if data_type == 'signinout':
            sign_in_out_data = None
        elif data_type == 'apply':
            apply_data = None
        elif data_type == 'otlieu':
            ot_lieu_data = None
        
        return jsonify({
            'success': True,
            'message': f'{data_type.capitalize()} data cleared successfully'
        })
        
    except Exception as e:
        return jsonify({'error': f'Error removing file: {str(e)}'}), 400

@app.route('/get_current_data_info', methods=['GET'])
def get_current_data_info():
    """Get information about currently loaded data"""
    data_type = request.args.get('data_type')
    
    if data_type == 'signinout':
        count = len(sign_in_out_data) if sign_in_out_data is not None else 0
        data = sign_in_out_data
    elif data_type == 'apply':
        count = len(apply_data) if apply_data is not None else 0
        data = apply_data
    elif data_type == 'otlieu':
        count = len(ot_lieu_data) if ot_lieu_data is not None else 0
        data = ot_lieu_data
    else:
        return jsonify({'error': 'Invalid data type'}), 400
    
    columns = list(data.columns) if data is not None else []
    
    return jsonify({
        'data_type': data_type,
        'row_count': count,
        'has_data': count > 0,
        'columns': columns
    })

@app.route('/save_apply_changes', methods=['POST'])
def save_apply_changes():
    """Save changes to Apply data"""
    global apply_data
    try:
        data = request.json.get('data', [])
        if data:
            df = pd.DataFrame(data)
            apply_data = df
            apply_data.to_excel(TEMP_APPLY_PATH, index=False)
            return jsonify({'success': True, 'message': 'Apply data saved successfully'})
        else:
            return jsonify({'success': False, 'error': 'No data provided'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/save_otlieu_changes', methods=['POST'])
def save_otlieu_changes():
    """Save changes to OT Lieu data"""
    global ot_lieu_data
    try:
        data = request.json.get('data', [])
        if data:
            df = pd.DataFrame(data)
            ot_lieu_data = df
            ot_lieu_save = ot_lieu_data.applymap(flatten_cell)
            ot_lieu_save.to_excel(TEMP_OTLIEU_PATH, index=False)
            return jsonify({'success': True, 'message': 'OT Lieu data saved successfully'})
        else:
            return jsonify({'success': False, 'error': 'No data provided'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

@app.route('/save_signinout_changes', methods=['POST'])
def save_signinout_changes():
    """Save changes to Sign In/Out data"""
    global sign_in_out_data
    try:
        data = request.json.get('data', [])
        if data:
            df = pd.DataFrame(data)
            sign_in_out_data = df
            sign_in_out_data.to_excel(TEMP_SIGNINOUT_PATH, index=False)
            return jsonify({'success': True, 'message': 'Sign In/Out data saved successfully'})
        else:
            return jsonify({'success': False, 'error': 'No data provided'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})

# ========================
# EXPORT ROUTES
# ========================
def calculate_otlieu_report_for_export():
    """Calculate OT Lieu Report data for export (without request context)"""
    global ot_lieu_data, employee_list_df, rules

    # Đảm bảo dữ liệu đầu vào
    if ot_lieu_data is None or ot_lieu_data.empty:
        temp_otlieu_path = os.path.join(app.config['UPLOAD_FOLDER'], 'temp_otlieu.xlsx')
        if os.path.exists(temp_otlieu_path):
            ot_lieu_data = pd.read_excel(temp_otlieu_path)
        else:
            return {'columns': [], 'rows': []}

    if employee_list_df is None or employee_list_df.empty:
        emp_path = os.path.join(app.config['UPLOAD_FOLDER'], 'employee_list.csv')
        if os.path.exists(emp_path):
            employee_list_df = pd.read_csv(emp_path, dtype=str)
        else:
            return {'columns': [], 'rows': []}

    if rules is None or rules.empty:
        rules_path = os.path.join(app.config['UPLOAD_FOLDER'], 'rules.xlsx')
        if os.path.exists(rules_path):
            rules = pd.read_excel(rules_path)
        else:
            return {'columns': [], 'rows': []}

    try:
        # Các cột cần thiết
        ot_payment_cols = [
            'OT Payment: Weekday Rate 150%',
            'OT Payment: Weekday-night Rate 200%',
            'OT Payment: Weekend Rate 200%',
            'OT Payment: Weekend-night Rate 270%',
            'OT Payment: Holiday Rate 300%',
            'OT Payment: Holiday-night Rate 390%',
        ]
        lieu_cols = [
            'Change in lieu: Weekday Rate 150%',
            'Change in lieu: Weekday-night Rate 200%',
            'Change in lieu: Weekend Rate 200%',
            'Change in lieu: Weekend-night Rate 270%',
            'Change in lieu: Holiday Rate 300%',
            'Change in lieu: Holiday-night Rate 390%',
        ]

        # Tính toán OT Lieu Before
        df = calculate_otlieu_before()
        if df is None or df.empty:
            df = pd.DataFrame(columns=['Name'] + ot_payment_cols + lieu_cols)

        # Lấy lieu remain previous month
        lieu_followup_path = os.path.join(app.config['UPLOAD_FOLDER'], 'lieu_followup.xlsx')
        lieu_remain_map = {}
        if os.path.exists(lieu_followup_path):
            lieu_followup_df = pd.read_excel(lieu_followup_path)
            if 'Name' in lieu_followup_df.columns and 'Lieu remain previous month' in lieu_followup_df.columns:
                for _, r in lieu_followup_df.iterrows():
                    name = str(r['Name']).strip()
                    try:
                        remain = float(r['Lieu remain previous month'])
                    except:
                        remain = 0.0
                    lieu_remain_map[name] = remain

        # Chuẩn hóa tên nhân viên - Extract ID from names for matching
        def extract_id_from_name(name_str):
            """Extract employee ID from name string"""
            import re
            if not isinstance(name_str, str):
                return ""
            # Look for 8+ digit numbers in the name
            match = re.search(r'(\d{8,})', name_str)
            return match.group(1) if match else name_str.strip()
        
        # Create mapping between OT data names and employee list names using ID
        df['EmployeeID'] = df['Name'].apply(extract_id_from_name)
        emp_list = employee_list_df.copy()
        emp_list['EmployeeID'] = emp_list['Name'].apply(extract_id_from_name)
        
        # Create ID to employee list name mapping
        id_to_emp_name = dict(zip(emp_list['EmployeeID'], emp_list['Name']))
        
        # Group theo employee ID thay vì Name
        grouped = df.groupby('EmployeeID')[ot_payment_cols + lieu_cols].sum().reset_index()

        # Build kết quả cho tất cả nhân viên
        result_rows = []
        for _, emp in emp_list.iterrows():
            emp_id_str = emp['EmployeeID']
            name = emp['Name']
            emp_id = emp.get('ID Number', '')
            row = {
                'No': int(emp.name) + 1,
                'Employee ID': emp_id,
                'Employee Name': name,
                'Lieu remain previous month': round(lieu_remain_map.get(name, 0.0), 3),
            }
            # Lấy dữ liệu OT/Lieu nếu có - sử dụng EmployeeID để match
            emp_data = grouped[grouped['EmployeeID'] == emp_id_str]
            for col in ot_payment_cols + lieu_cols:
                value = round(float(emp_data[col].values[0]), 3) if not emp_data.empty and col in emp_data.columns else 0.0
                row[col] = '' if value == 0.0 else value

            # Tổng hợp
            total_ot_paid = sum([row[c] if row[c] != '' else 0.0 for c in ot_payment_cols])
            total_used_hours = sum([row[c] if row[c] != '' else 0.0 for c in lieu_cols])
            row['Total OT paid'] = '' if total_ot_paid == 0.0 else round(total_ot_paid, 3)
            row['Total used hours in month'] = '' if total_used_hours == 0.0 else round(total_used_hours, 3)
            
            lieu_remain = round(row['Lieu remain previous month'] - total_used_hours, 3)
            row['Remain unused time off in lieu'] = '' if lieu_remain == 0.0 else lieu_remain
            
            # OVERTIME (For payment)
            row['OVERTIME (For payment)'] = '' if total_ot_paid == 0.0 else total_ot_paid
            # OVERTIME (No pay, For later time in lieu)
            row['OVERTIME (No pay, For later time in lieu)'] = ''  # Default empty
            # Time off in lieu (hour)
            row['Time off in lieu (hour)'] = '' if total_used_hours == 0.0 else total_used_hours
            # Transferred to normal working hours
            if total_used_hours > 0 and total_ot_paid == 0:
                row['Transferred to normal working hours'] = total_used_hours
            elif total_ot_paid > 25:
                row['Transferred to normal working hours'] = round(total_ot_paid - 25, 3)
            else:
                row['Transferred to normal working hours'] = ''
            row['Date'] = ''

            # Handle lieu remain previous month display
            if row['Lieu remain previous month'] == 0.0:
                row['Lieu remain previous month'] = ''

            result_rows.append(row)

        # Đảm bảo đúng thứ tự cột
        col_order = [
            'No', 'Employee ID', 'Employee Name',
            'OVERTIME (For payment)', 'OVERTIME (No pay, For later time in lieu)', 'Time off in lieu (hour)',
            'Remain unused time off in lieu', 'Total OT paid',
            'OT Payment: Weekday Rate 150%', 'OT Payment: Weekday-night Rate 200%',
            'OT Payment: Weekend Rate 200%', 'OT Payment: Weekend-night Rate 270%',
            'OT Payment: Holiday Rate 300%', 'OT Payment: Holiday-night Rate 390%',
            'Change in lieu: Weekday Rate 150%', 'Change in lieu: Weekday-night Rate 200%',
            'Change in lieu: Weekend Rate 200%', 'Change in lieu: Weekend-night Rate 270%',
            'Change in lieu: Holiday Rate 300%', 'Change in lieu: Holiday-night Rate 390%',
            'Transferred to normal working hours', 'Date', 'Total used hours in month'
        ]
        # Đổi tên cột cho thân thiện
        col_rename = {
            'OT Payment: Weekday Rate 150%': 'OT weekday 150%',
            'OT Payment: Weekday-night Rate 200%': 'OT weekday night 200%',
            'OT Payment: Weekend Rate 200%': 'OT weekly holiday 200%',
            'OT Payment: Weekend-night Rate 270%': 'OT weekly holiday night 270%',
            'OT Payment: Holiday Rate 300%': 'OT public holiday 300%',
            'OT Payment: Holiday-night Rate 390%': 'OT public holiday night 390%',
            'Change in lieu: Weekday Rate 150%': 'OT weekday 150% (lieu)',
            'Change in lieu: Weekday-night Rate 200%': 'OT weekday night 200% (lieu)',
            'Change in lieu: Weekend Rate 200%': 'OT weekly holiday 200% (lieu)',
            'Change in lieu: Weekend-night Rate 270%': 'OT weekly holiday night 270% (lieu)',
            'Change in lieu: Holiday Rate 300%': 'OT public holiday 300% (lieu)',
            'Change in lieu: Holiday-night Rate 390%': 'OT public holiday night 390% (lieu)',
        }
        result_df = pd.DataFrame(result_rows)
        result_df = result_df.rename(columns=col_rename)
        col_order = [
            'No', 'Employee ID', 'Employee Name',
            'OVERTIME (For payment)', 'OVERTIME (No pay, For later time in lieu)', 'Time off in lieu (hour)',
            'Remain unused time off in lieu', 'Total OT paid',
            'OT weekday 150%', 'OT weekday night 200%', 'OT weekly holiday 200%',
            'OT weekly holiday night 270%', 'OT public holiday 300%', 'OT public holiday night 390%',
            'OT weekday 150% (lieu)', 'OT weekday night 200% (lieu)', 'OT weekly holiday 200% (lieu)',
            'OT weekly holiday night 270% (lieu)', 'OT public holiday 300% (lieu)', 'OT public holiday night 390% (lieu)',
            'Transferred to normal working hours', 'Date', 'Total used hours in month'
        ]
        result_df = result_df[[c for c in col_order if c in result_df.columns]]

        cols = list(result_df.columns)
        rows = result_df.fillna('').astype(str).values.tolist()
        return {'columns': cols, 'rows': rows}

    except Exception as e:
        print(f"Error in calculate_otlieu_report_for_export: {e}")
        return {'columns': [], 'rows': []}

def calculate_total_attendance_detail_for_export(month=None, year=None, employee_list_df=None, apply_data=None, ot_lieu_data=None, sign_in_out_data=None):
    """Calculate Total Attendance Detail data for export - uses same logic as frontend"""
    try:
        # Call get_total_attendance_detail to get the same data as frontend
        with app.test_request_context():
            response = get_total_attendance_detail()
            
        if response.status_code == 200:
            data = response.get_json()
            if data and 'columns' in data and 'data' in data:
                return {
                    'columns': data['columns'],
                    'rows': data['data']  # Frontend uses 'data' key, not 'rows'
                }
        
        # Return empty data if no response
        return {'columns': [], 'rows': []}
        
    except Exception as e:
        print(f"Error in calculate_total_attendance_detail_for_export: {e}")
        return {'columns': [], 'rows': []}


def calculate_abnormal_late_early_for_export(month=None, year=None):
    """Calculate Abnormal Late/Early data for export - uses same logic as frontend"""
    try:
        # Call get_attendance_report to get the same data as frontend
        with app.test_request_context():
            response = get_attendance_report()
            
        if response.status_code == 200:
            data = response.get_json()
            if data and 'abnormal_late_early_columns' in data and 'abnormal_late_early_rows' in data:
                return {
                    'columns': data['abnormal_late_early_columns'],
                    'rows': data['abnormal_late_early_rows']
                }
        
        # Return empty data if no response
        return {'columns': [], 'rows': []}
        
    except Exception as e:
        print(f"Error in calculate_abnormal_late_early_for_export: {e}")
        return {'columns': [], 'rows': []}

def calculate_abnormal_missing_for_export(month=None, year=None):
    """Calculate Abnormal Missing data for export"""
    global sign_in_out_data, apply_data, employee_list_df, rules
    
    try:
        # Load data from temp files if global variables are empty
        if sign_in_out_data is None or sign_in_out_data.empty:
            temp_signinout_path = os.path.join(app.config['UPLOAD_FOLDER'], 'temp_signinout.xlsx')
            if os.path.exists(temp_signinout_path):
                sign_in_out_data = pd.read_excel(temp_signinout_path)
            else:
                return pd.DataFrame()
        
        if employee_list_df is None or employee_list_df.empty:
            emp_path = os.path.join(app.config['UPLOAD_FOLDER'], 'employee_list.csv')
            if os.path.exists(emp_path):
                employee_list_df = pd.read_csv(emp_path, dtype=str)
            else:
                return pd.DataFrame()

        if apply_data is None or apply_data.empty:
            temp_apply_path = os.path.join(app.config['UPLOAD_FOLDER'], 'temp_apply.xlsx')
            if os.path.exists(temp_apply_path):
                apply_data = pd.read_excel(temp_apply_path)

        # Determine month and year
        if month and year:
            target_month = month
            target_year = year
            print(f"calculate_abnormal_missing_for_export: Using provided month/year: {month}/{year}")
        else:
            # Auto-detect from data using most common month (same logic as get_attendance_report)
            target_month, target_year = 7, 2024  # Default fallback
            if sign_in_out_data is not None and not sign_in_out_data.empty:
                dates = [pd.to_datetime(r.get('attendance_time')) for _, r in sign_in_out_data.iterrows() if pd.notna(r.get('attendance_time'))]
                if dates:
                    month_counts = {}
                    for date in dates:
                        month_key = (date.month, date.year)
                        month_counts[month_key] = month_counts.get(month_key, 0) + 1
                    most_common_month = max(month_counts.items(), key=lambda x: x[1])[0]
                    target_month, target_year = most_common_month
            print(f"calculate_abnormal_missing_for_export: Auto-detected month/year: {target_month}/{target_year}")

        # Calculate date range: 19th of previous month to 20th of current month
        if target_month == 1:
            prev_month = 12
            prev_year = target_year - 1
        else:
            prev_month = target_month - 1
            prev_year = target_year
        
        start_date = pd.Timestamp(prev_year, prev_month, 19)
        end_date = pd.Timestamp(target_year, target_month, 20)
        date_range = pd.date_range(start=start_date, end=end_date, freq='D')
        
        # Create abnormal missing data
        missing_data = []
        
        # Extract names from employee list for filtering
        valid_names = set(employee_list_df['Name'].astype(str)) if not employee_list_df.empty else set()
        
        # Convert date_range to date objects for comparison
        valid_dates = set(d.date() for d in date_range)
        
        # Get sign-in/out data for processing
        if not sign_in_out_data.empty and 'attendance_time' in sign_in_out_data.columns:
            df_sign = sign_in_out_data.copy()
            df_sign['attendance_time'] = pd.to_datetime(df_sign['attendance_time'], errors='coerce')
            df_sign['date'] = df_sign['attendance_time'].dt.date
            
            # Check for missing attendance for each employee on each workday
            for name in valid_names:
                emp_records = df_sign[df_sign['emp_name'].astype(str) == name]
                emp_dates = set(emp_records['date'].dropna())
                
                for date in date_range:
                    # Skip weekends (Saturday=5, Sunday=6) unless it's a special workday
                    if date.weekday() >= 5:
                        continue
                        
                    check_date = date.date()
                    
                    # Check if employee has attendance record for this date
                    if check_date not in emp_dates:
                        # Check if employee has approved leave for this date
                        has_leave = False
                        if apply_data is not None and not apply_data.empty:
                            for _, leave_record in apply_data.iterrows():
                                if (leave_record.get('Emp Name', '').strip() == name and 
                                    leave_record.get('Results', '').strip() == 'Approved'):
                                    try:
                                        start_date = pd.to_datetime(leave_record.get('Start Date')).date()
                                        end_date = pd.to_datetime(leave_record.get('End Date')).date()
                                        if start_date <= check_date <= end_date:
                                            has_leave = True
                                            break
                                    except:
                                        continue
                        
                        # If no leave, this is missing attendance
                        if not has_leave:
                            missing_data.append({
                                'Employee': name,
                                'Date': check_date.strftime('%Y-%m-%d'),
                                'SignIn': '',
                                'SignOut': '',
                                'Status': 'Missing',
                                'Reason': 'No attendance record found',
                                'HasLeave': 'No'
                            })
        
        if not missing_data:
            return pd.DataFrame()
        
        return pd.DataFrame(missing_data)
        
    except Exception as e:
        print(f"Error in calculate_abnormal_missing_for_export: {e}")
        return pd.DataFrame()

if __name__ == '__main__':
    import os
    port = int(os.environ.get('PORT', 5000))
    debug = os.environ.get('FLASK_ENV', 'production') == 'development'
    app.run(debug=debug, host='0.0.0.0', port=port) 